"""
Saidas da analise de implantacao — Excel e PDF
Sigma Contabilidade

Tres entregaveis:
  1. Mapa de rubricas   — origem -> grupo canonico -> codigo no sistema de destino
  2. Planilha de import — verbas por funcionario x competencia
  3. Ficha do funcionario (PDF) — historico mes a mes para conferir na implantacao
"""
import io
import re
from datetime import datetime

from app.core.textutils import competencia_label, fmt_cpf
from app.services import rubrics

SIGMA_BORDO = "A72C31"
CINZA = "F5F5F5"

CLASSE_LABEL = {
    "fixa": "Fixa",
    "fixa_reajustada": "Fixa (com reajuste)",
    "variavel": "Variavel",
    "eventual": "Eventual",
    "calculada": "Calculada pelo sistema",
}

CLASSE_ACAO = {
    "fixa": "Cadastrar como evento FIXO no funcionario",
    "fixa_reajustada": "Cadastrar como evento FIXO com o valor do ultimo mes",
    "variavel": "Lancar mes a mes (evento variavel)",
    "eventual": "Lancar apenas quando ocorrer",
    "calculada": "Nao importar — o sistema recalcula",
}


# ─────────────────────────────────────────────
# HELPERS DE ESTILO
# ─────────────────────────────────────────────

def _estilo_cabecalho(ws, linha, n_colunas):
    from openpyxl.styles import Font, PatternFill, Alignment
    for c in range(1, n_colunas + 1):
        cel = ws.cell(row=linha, column=c)
        cel.font = Font(bold=True, color="FFFFFF", size=10)
        cel.fill = PatternFill("solid", fgColor=SIGMA_BORDO)
        cel.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[linha].height = 28


def _ajusta_larguras(ws, larguras):
    from openpyxl.utils import get_column_letter
    for i, w in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _moeda(ws, col_letras, primeira_linha, ultima_linha):
    for letra in col_letras:
        for r in range(primeira_linha, ultima_linha + 1):
            ws[f"{letra}{r}"].number_format = 'R$ #,##0.00'


# ─────────────────────────────────────────────
# 1. MAPA DE RUBRICAS
# ─────────────────────────────────────────────

def mapa_rubricas(analise: dict) -> bytes:
    """
    Planilha de-para: rubrica do escritorio anterior -> codigo no sistema da Sigma.

    A coluna 'Codigo destino' vem preenchida quando o grupo ja tem
    'codigo_destino' no rubricas-equivalentes.json; o resto fica em branco
    para decisao — inventar codigo de destino seria pior que deixar vazio.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mapa de Rubricas"

    resumo = analise.get("resumo", {})
    ws["A1"] = "MAPA DE RUBRICAS — IMPLANTACAO DE CLIENTE NOVO"
    ws["A1"].font = Font(bold=True, size=13, color=SIGMA_BORDO)
    ws["A2"] = f"{resumo.get('empresa', {}).get('nome', '')}   |   Periodo: {resumo.get('periodo', '-')}"
    ws["A3"] = f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')} — Sigma Contabilidade"
    ws["A3"].font = Font(size=8, color="666666")

    cabecalho = [
        "Rubrica (origem)", "Outras grafias", "Cod. origem", "Tipo",
        "Grupo Sigma", "Codigo destino", "Classificacao", "O que fazer",
        "Funcionarios", "Meses", "Menor valor", "Maior valor", "Total no periodo",
    ]
    ws.append([])
    ws.append(cabecalho)
    linha_cab = ws.max_row
    _estilo_cabecalho(ws, linha_cab, len(cabecalho))

    for c in analise.get("catalogo_rubricas", []):
        grupo = c.get("grupo", "")
        cod_destino = _codigo_destino(grupo)

        outras = [d for d in c.get("descricoes_vistas", []) if d != c["descricao"]]
        ws.append([
            c["descricao"],
            " | ".join(outras),
            ", ".join(c.get("codigos_origem", [])),
            {"provento": "Provento", "desconto": "Desconto"}.get(c["tipo"], c["tipo"]),
            grupo if grupo in rubrics.RUBRIC_META else "(sem equivalencia)",
            cod_destino,
            CLASSE_LABEL.get(c["classe"], c["classe"]),
            CLASSE_ACAO.get(c["classe"], ""),
            c["funcionarios"],
            c["ocorrencias"],
            c["min"], c["max"], c["total"],
        ])

    ultima = ws.max_row
    _moeda(ws, ["K", "L", "M"], linha_cab + 1, ultima)
    _ajusta_larguras(ws, [34, 30, 12, 10, 22, 14, 20, 42, 12, 8, 14, 14, 16])
    ws.freeze_panes = ws.cell(row=linha_cab + 1, column=1)

    # Destaca o que exige decisao
    amarelo = PatternFill("solid", fgColor="FFF3CD")
    for r in range(linha_cab + 1, ultima + 1):
        if ws[f"E{r}"].value == "(sem equivalencia)" or not ws[f"F{r}"].value:
            for col in "ABCDEFGH":
                ws[f"{col}{r}"].fill = amarelo

    ws2 = wb.create_sheet("Como usar")
    for i, txt in enumerate([
        "COMO USAR ESTE MAPA",
        "",
        "1. As linhas em amarelo precisam de decisao: ou nao tem equivalencia no dicionario",
        "   da Sigma, ou ainda nao tem codigo de destino definido.",
        "2. Preencha a coluna 'Codigo destino' com o codigo da rubrica no sistema da Sigma.",
        "3. Depois de preencher, leve os codigos para o arquivo rubricas-equivalentes.json",
        "   (campo 'codigo_destino' de cada grupo) — assim o proximo cliente ja sai mapeado.",
        "",
        "CLASSIFICACAO:",
        "  Fixa                   -> mesmo valor todos os meses. Cadastrar como evento fixo.",
        "  Fixa (com reajuste)    -> valor constante que mudou de patamar. Use o valor do ultimo mes.",
        "  Variavel               -> muda todo mes. Lancar mes a mes.",
        "  Eventual               -> aparece so em alguns meses. Lancar quando ocorrer.",
        "  Calculada pelo sistema -> INSS, IRRF, FGTS. Nao importar: o destino recalcula.",
    ], start=1):
        ws2.cell(row=i, column=1, value=txt)
    ws2.cell(row=1, column=1).font = Font(bold=True, size=12, color=SIGMA_BORDO)
    _ajusta_larguras(ws2, [100])

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _codigo_destino(grupo: str) -> str:
    """Le o codigo de destino cadastrado no rubricas-equivalentes.json."""
    if not grupo:
        return ""
    try:
        import json
        import os
        p = os.path.join(
            os.path.dirname(os.path.dirname(os.path.dirname(__file__))),
            "rubricas-equivalentes.json",
        )
        with open(p, encoding="utf-8") as f:
            d = json.load(f)
        return d.get("grupos", {}).get(grupo, {}).get("codigo_destino", "") or ""
    except Exception:
        return ""


# ─────────────────────────────────────────────
# 2. PLANILHA DE IMPORTACAO
# ─────────────────────────────────────────────

def planilha_importacao(analise: dict) -> bytes:
    """
    Verbas por funcionario x competencia, prontas para conferencia e digitacao.

    Aba 1 — Eventos fixos    (cadastrar uma vez no cadastro do funcionario)
    Aba 2 — Eventos variaveis (lancar mes a mes)
    Aba 3 — Cadastro         (dados do funcionario)
    Aba 4 — Resumo mensal    (totais por competencia, para bater com a folha antiga)
    """
    import openpyxl
    from openpyxl.styles import Font

    wb = openpyxl.Workbook()
    comps = analise.get("competencias", [])
    funcs = analise.get("funcionarios", [])

    # ── Aba Cadastro ────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Cadastro"
    cab = ["Nome", "CPF", "Matricula origem", "CBO", "Funcao", "Admissao",
           "1o mes na folha", "Ultimo mes", "Salario base (ultimo)", "Situacao"]
    ws.append(cab)
    _estilo_cabecalho(ws, 1, len(cab))
    for f in funcs:
        ativos = f.get("meses_ativos", [])
        sal = None
        for c in reversed(ativos):
            if f["meses"][c].get("salario_base"):
                sal = f["meses"][c]["salario_base"]
                break
        saiu = bool(ativos) and comps and ativos[-1] != comps[-1]
        ws.append([
            f["nome"], fmt_cpf(f.get("cpf")), f.get("matricula", ""), f.get("cbo", ""),
            f.get("funcao", ""), f.get("admissao", ""),
            competencia_label(ativos[0]) if ativos else "-",
            competencia_label(ativos[-1]) if ativos else "-",
            sal, "Desligado no periodo" if saiu else "Ativo",
        ])
    _moeda(ws, ["I"], 2, ws.max_row)
    _ajusta_larguras(ws, [30, 16, 14, 10, 26, 12, 15, 15, 18, 20])
    ws.freeze_panes = "A2"

    # ── Aba Eventos fixos ───────────────────────────────────────────
    wf = wb.create_sheet("Eventos fixos")
    cab = ["Nome", "CPF", "Rubrica (origem)", "Grupo Sigma", "Codigo destino",
           "Tipo", "Valor a cadastrar", "Classificacao", "Meses observados"]
    wf.append(cab)
    _estilo_cabecalho(wf, 1, len(cab))
    for f in funcs:
        for kr, info in sorted(f.get("rubricas", {}).items()):
            if info["classe"] not in ("fixa", "fixa_reajustada"):
                continue
            wf.append([
                f["nome"], fmt_cpf(f.get("cpf")), info.get("descricao", kr),
                info.get("grupo", ""), _codigo_destino(info.get("grupo", "")),
                "Provento" if info.get("tipo") == "provento" else "Desconto",
                info["ultimo"], CLASSE_LABEL.get(info["classe"], info["classe"]),
                f"{info['meses_presente']}/{info['meses_ativos']}",
            ])
    _moeda(wf, ["G"], 2, wf.max_row)
    _ajusta_larguras(wf, [30, 16, 32, 20, 14, 10, 18, 20, 16])
    wf.freeze_panes = "A2"

    # ── Aba Eventos variaveis ───────────────────────────────────────
    wv = wb.create_sheet("Eventos variaveis")
    cab = ["Nome", "CPF", "Rubrica (origem)", "Grupo Sigma", "Codigo destino", "Tipo"] \
        + [competencia_label(c) for c in comps] + ["Media", "Total"]
    wv.append(cab)
    _estilo_cabecalho(wv, 1, len(cab))
    for f in funcs:
        for kr, info in sorted(f.get("rubricas", {}).items()):
            if info["classe"] not in ("variavel", "eventual"):
                continue
            linha = [
                f["nome"], fmt_cpf(f.get("cpf")), info.get("descricao", kr),
                info.get("grupo", ""), _codigo_destino(info.get("grupo", "")),
                "Provento" if info.get("tipo") == "provento" else "Desconto",
            ]
            linha += [info["valores"].get(c, None) for c in comps]
            linha += [info["media"], round(sum(info["valores"].values()), 2)]
            wv.append(linha)
    n_fixas = 6
    from openpyxl.utils import get_column_letter
    cols_valor = [get_column_letter(i) for i in range(n_fixas + 1, n_fixas + len(comps) + 3)]
    _moeda(wv, cols_valor, 2, wv.max_row)
    _ajusta_larguras(wv, [30, 16, 32, 20, 14, 10] + [14] * len(comps) + [14, 14])
    wv.freeze_panes = "C2"

    # ── Aba Resumo mensal ───────────────────────────────────────────
    wr = wb.create_sheet("Resumo mensal")
    cab = ["Competencia", "Funcionarios", "Total proventos", "Total descontos", "Liquido"]
    wr.append(cab)
    _estilo_cabecalho(wr, 1, len(cab))
    for p in analise.get("resumo", {}).get("por_competencia", []):
        wr.append([p["label"], p["funcionarios"], p["proventos"], p["descontos"], p["liquido"]])
    _moeda(wr, ["C", "D", "E"], 2, wr.max_row)
    _ajusta_larguras(wr, [18, 14, 18, 18, 18])
    wr["A" + str(wr.max_row + 2)] = (
        "Confira estes totais contra a folha do escritorio anterior antes de importar."
    )
    wr["A" + str(wr.max_row)].font = Font(italic=True, size=9, color="666666")

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ─────────────────────────────────────────────
# 3. FICHA POR FUNCIONARIO (PDF)
# ─────────────────────────────────────────────

def fichas_funcionarios(analise: dict) -> bytes:
    """Uma pagina por funcionario com o historico completo mes a mes."""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak,
    )

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=12 * mm, rightMargin=12 * mm,
        topMargin=12 * mm, bottomMargin=12 * mm,
        title="Fichas de Implantacao — Sigma Contabilidade",
    )
    ss = getSampleStyleSheet()
    h1 = ParagraphStyle("h1", parent=ss["Heading1"], fontSize=14,
                        textColor=colors.HexColor("#" + SIGMA_BORDO), spaceAfter=2)
    sub = ParagraphStyle("sub", parent=ss["Normal"], fontSize=8,
                         textColor=colors.HexColor("#666666"), spaceAfter=6)
    h2 = ParagraphStyle("h2", parent=ss["Heading2"], fontSize=10,
                        textColor=colors.HexColor("#333333"), spaceBefore=8, spaceAfter=3)
    small = ParagraphStyle("small", parent=ss["Normal"], fontSize=7.5)

    comps = analise.get("competencias", [])
    resumo = analise.get("resumo", {})
    story = []

    def brl(v):
        if v is None:
            return "-"
        return f"{v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

    for idx, f in enumerate(analise.get("funcionarios", [])):
        if idx:
            story.append(PageBreak())

        story.append(Paragraph(f["nome"], h1))
        ident = " &nbsp;|&nbsp; ".join(x for x in [
            f"CPF {fmt_cpf(f['cpf'])}" if f.get("cpf") else "",
            f"Matricula {f['matricula']}" if f.get("matricula") else "",
            f"CBO {f['cbo']}" if f.get("cbo") else "",
            f["funcao"] if f.get("funcao") else "",
            f"Admissao {f['admissao']}" if f.get("admissao") else "",
        ] if x)
        story.append(Paragraph(ident or "-", sub))
        story.append(Paragraph(
            f"{resumo.get('empresa', {}).get('nome', '')} — periodo analisado: "
            f"{resumo.get('periodo', '-')}", sub))

        ativos = f.get("meses_ativos", [])

        # Tabela de rubricas x competencias
        story.append(Paragraph("Rubricas mes a mes", h2))
        cab = ["Rubrica", "Tipo", "Classe"] + [competencia_label(c)[:3] + "/" + c[2:4] for c in comps]
        dados = [cab]
        for kr, info in sorted(
            f.get("rubricas", {}).items(),
            key=lambda kv: (kv[1].get("tipo") != "provento", -sum(kv[1]["valores"].values())),
        ):
            linha = [
                Paragraph(info.get("descricao", kr)[:38], small),
                "P" if info.get("tipo") == "provento" else "D",
                CLASSE_LABEL.get(info["classe"], info["classe"])[:12],
            ]
            linha += [brl(info["valores"].get(c)) for c in comps]
            dados.append(linha)

        tot = ["TOTAL LIQUIDO", "", ""]
        for c in comps:
            tot.append(brl(f["meses"][c]["liquido"]) if c in f["meses"] else "-")
        dados.append(tot)

        # Largura das colunas de competencia: divide o espaco restante, com teto
        larg = [58 * mm, 8 * mm, 22 * mm] + [
            min(26, 184 / max(len(comps), 1)) * mm for _ in comps
        ]
        t = Table(dados, colWidths=larg, repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#" + SIGMA_BORDO)),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("ALIGN", (3, 0), (-1, -1), "RIGHT"),
            ("ALIGN", (1, 0), (2, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#DDDDDD")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#" + CINZA)]),
            ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#EDEDED")),
            ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
            ("TOPPADDING", (0, 0), (-1, -1), 2),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ]))
        story.append(t)

        # Eventos
        eventos = f.get("eventos", [])
        if eventos:
            story.append(Paragraph("O que mudou no periodo", h2))
            de = [["Competencia", "Ocorrencia", "Detalhe"]]
            for e in eventos:
                de.append([
                    competencia_label(e["competencia"]),
                    Paragraph(e["titulo"], small),
                    Paragraph(e["detalhe"], small),
                ])
            te = Table(de, colWidths=[26 * mm, 78 * mm, 160 * mm], repeatRows=1)
            te.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#333333")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#DDDDDD")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]))
            story.append(te)

        if not ativos:
            story.append(Paragraph("Sem recibos identificados no periodo.", small))

    if not story:
        story = [Paragraph("Nenhum funcionario analisado.", ss["Normal"])]

    doc.build(story)
    return buf.getvalue()


def nome_arquivo(analise: dict, prefixo: str, ext: str) -> str:
    """Nome de arquivo previsivel: prefixo_empresa_periodo.ext"""
    emp = analise.get("resumo", {}).get("empresa", {}).get("nome", "cliente")
    emp = re.sub(r"[^\w]+", "-", emp).strip("-").lower()[:32] or "cliente"
    comps = analise.get("competencias", [])
    per = f"{comps[0]}_a_{comps[-1]}" if comps else datetime.now().strftime("%Y-%m")
    return f"{prefixo}_{emp}_{per}.{ext}"
