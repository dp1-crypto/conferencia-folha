"""
Exportacao para Excel e CSV
"""
import io
import csv
from app.core.utils import fmt_brl


def exportar_excel(dados: dict, tipo: str) -> bytes:
    """
    tipo: 'folha' | 'beneficio' | 'mes_anterior' | 'impostos'
    Gera .xlsx usando openpyxl.
    """
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import datetime

    wb = openpyxl.Workbook()

    # Estilos
    red_fill = PatternFill(start_color='A72C31', end_color='A72C31', fill_type='solid')
    light_red = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFFBEB', end_color='FFFBEB', fill_type='solid')
    green_fill = PatternFill(start_color='DCFCE7', end_color='DCFCE7', fill_type='solid')
    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    title_font = Font(name='Calibri', bold=True, color='A72C31', size=13)
    bold_font = Font(name='Calibri', bold=True, size=10)

    def _header_row(ws, cols):
        ws.append(cols)
        for cell in ws[ws.max_row]:
            cell.fill = red_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

    def _auto_width(ws):
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    max_len = max(max_len, len(str(cell.value or '')))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

    if tipo == 'folha':
        ws = wb.active
        ws.title = 'Divergencias'
        ws['A1'] = 'Conferencia de Folha -- Sigma Assessoria'
        ws['A1'].font = title_font
        ws['A2'] = f'Gerado em: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
        ws.append([])

        resumo = dados.get('resumo', {})
        ws.append(['RESUMO EXECUTIVO'])
        ws.append(['Total de colaboradores', resumo.get('total', 0)])
        ws.append(['Colaboradores OK', resumo.get('ok', 0)])
        ws.append(['Com divergencias', resumo.get('divergencias', 0)])
        ws.append([])

        _header_row(ws, ['Colaborador', 'Tipo de Divergencia', 'Descricao', 'Criticidade'])

        for emp in dados.get('funcionarios', []):
            if emp.get('divs'):
                for d in emp['divs']:
                    row = [
                        emp.get('nome_exibir', emp.get('nome', '')),
                        d.get('tipo', ''),
                        d.get('desc', ''),
                        d.get('g', '').upper(),
                    ]
                    ws.append(row)
                    last_row = ws[ws.max_row]
                    g = d.get('g', '')
                    if g == 'alta':
                        for cell in last_row: cell.fill = light_red
                    elif g == 'media':
                        for cell in last_row: cell.fill = yellow_fill

        _auto_width(ws)

        # Aba todos colaboradores
        ws2 = wb.create_sheet('Todos os Colaboradores')
        _header_row(ws2, ['Colaborador', 'Status', 'Qtd Divergencias'])
        for emp in dados.get('funcionarios', []):
            ws2.append([
                emp.get('nome_exibir', emp.get('nome', '')),
                emp.get('status', 'OK'),
                len(emp.get('divs', [])),
            ])
        _auto_width(ws2)

    elif tipo == 'beneficio':
        ws = wb.active
        ws.title = 'Beneficios'
        ws['A1'] = 'Conferencia de Beneficios -- Sigma Assessoria'
        ws['A1'].font = title_font
        ws.append([])

        ws.append(['Total esperado', fmt_brl(dados.get('total_esperado', 0))])
        ws.append(['Total descontado', fmt_brl(dados.get('total_extrato', 0))])
        ws.append(['Diferenca', fmt_brl(dados.get('total_diferenca', 0))])
        ws.append([])

        _header_row(ws, ['Colaborador', 'Mensalidade', 'Valor Esperado', 'Descontado', 'Diferenca', 'Status'])
        for r in dados.get('resultados', []):
            ws.append([
                r.get('nome', ''),
                r.get('total_fatura', 0),
                r.get('valor_esperado', 0),
                r.get('valor_descontado', 0),
                r.get('diferenca', 0),
                r.get('status', ''),
            ])
            if r.get('status') == 'MAIOR':
                for cell in ws[ws.max_row]: cell.fill = light_red
            elif r.get('status') == 'MENOR':
                for cell in ws[ws.max_row]: cell.fill = yellow_fill
        _auto_width(ws)

    elif tipo == 'mes_anterior':
        ws = wb.active
        ws.title = 'Comparativo Mensal'
        ws['A1'] = 'Conferencia Mes Anterior -- Sigma Assessoria'
        ws['A1'].font = title_font
        ws.append([])

        resumo = dados.get('resumo', {})
        ws.append(['Total atual', resumo.get('total_atual', 0)])
        ws.append(['Colaboradores novos', resumo.get('novos', 0)])
        ws.append(['Desligados', resumo.get('desligados', 0)])
        ws.append(['Com alteracoes', resumo.get('alterados', 0)])
        ws.append([])

        _header_row(ws, ['Colaborador', 'Campo', 'Valor Anterior', 'Valor Atual', 'Diferenca', 'Variacao %', 'Criticidade'])
        for alt in dados.get('alteracoes', []):
            ws.append([
                alt.get('nome', ''),
                alt.get('campo', ''),
                alt.get('valor_anterior', 0),
                alt.get('valor_atual', 0),
                alt.get('diferenca', 0),
                f"{alt.get('pct_variacao', 0):.1f}%",
                alt.get('criticidade', '').upper(),
            ])
            if alt.get('criticidade') == 'alta':
                for cell in ws[ws.max_row]: cell.fill = light_red
            elif alt.get('criticidade') == 'media':
                for cell in ws[ws.max_row]: cell.fill = yellow_fill

        if dados.get('colaboradores_novos'):
            ws.append([])
            ws.append(['NOVOS COLABORADORES'])
            _header_row(ws, ['Nome', 'Liquido'])
            for n in dados['colaboradores_novos']:
                ws.append([n.get('nome', ''), n.get('liquido', 0)])
        _auto_width(ws)

    elif tipo == 'impostos':
        ws = wb.active
        ws.title = 'Auditoria INSS IRRF'
        ws['A1'] = 'Auditoria de Impostos -- Sigma Assessoria'
        ws['A1'].font = title_font
        ws.append([])

        resumo = dados.get('resumo', {})
        ws.append(['Total colaboradores', resumo.get('total', 0)])
        ws.append(['Com divergencia', resumo.get('com_divergencia', 0)])
        ws.append(['OK', resumo.get('ok', 0)])
        ws.append([])

        _header_row(ws, ['Colaborador', 'Salario Bruto', 'INSS Calculado', 'INSS Encontrado', 'Status INSS',
                         'Base IRRF', 'IRRF Calculado', 'IRRF Encontrado', 'Status IRRF'])
        for c in dados.get('colaboradores', []):
            ws.append([
                c.get('nome', ''),
                c.get('salario_bruto', 0),
                c.get('inss_calculado', 0),
                c.get('inss_encontrado', 0),
                c.get('inss_status', ''),
                c.get('base_irrf', 0),
                c.get('irrf_calculado', 0),
                c.get('irrf_encontrado', 0),
                c.get('irrf_status', ''),
            ])
            if c.get('divergencias'):
                for cell in ws[ws.max_row]: cell.fill = light_red
            elif c.get('inss_status') == 'OK' and c.get('irrf_status') == 'OK':
                for cell in ws[ws.max_row]: cell.fill = green_fill
        _auto_width(ws)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def exportar_csv(dados: dict, tipo: str) -> str:
    """Exporta CSV das divergencias."""
    output = io.StringIO()

    if tipo == 'folha':
        writer = csv.writer(output)
        writer.writerow(['Colaborador', 'Tipo', 'Descricao', 'Criticidade'])
        for emp in dados.get('funcionarios', []):
            for d in emp.get('divs', []):
                writer.writerow([
                    emp.get('nome_exibir', emp.get('nome', '')),
                    d.get('tipo', ''),
                    d.get('desc', ''),
                    d.get('g', '').upper(),
                ])
    elif tipo == 'beneficio':
        writer = csv.writer(output)
        writer.writerow(['Colaborador', 'Mensalidade', 'Esperado', 'Descontado', 'Diferenca', 'Status'])
        for r in dados.get('resultados', []):
            writer.writerow([r.get('nome',''), r.get('total_fatura',0), r.get('valor_esperado',0),
                           r.get('valor_descontado',0), r.get('diferenca',0), r.get('status','')])
    elif tipo == 'mes_anterior':
        writer = csv.writer(output)
        writer.writerow(['Colaborador', 'Campo', 'Anterior', 'Atual', 'Diferenca', 'Variacao%', 'Criticidade'])
        for a in dados.get('alteracoes', []):
            writer.writerow([a.get('nome',''), a.get('campo',''), a.get('valor_anterior',0),
                           a.get('valor_atual',0), a.get('diferenca',0), a.get('pct_variacao',0), a.get('criticidade','')])
    elif tipo == 'impostos':
        writer = csv.writer(output)
        writer.writerow(['Colaborador', 'Salario Bruto', 'INSS Calc', 'INSS Enc', 'Status INSS', 'IRRF Calc', 'IRRF Enc', 'Status IRRF'])
        for c in dados.get('colaboradores', []):
            writer.writerow([c.get('nome',''), c.get('salario_bruto',0), c.get('inss_calculado',0),
                           c.get('inss_encontrado',0), c.get('inss_status',''), c.get('irrf_calculado',0),
                           c.get('irrf_encontrado',0), c.get('irrf_status','')])

    return output.getvalue()
