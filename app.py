#!/usr/bin/env python3
"""
Sigma Contabilidade — Conferência de Folha de Pagamento
Compara Word (instruções) + Excel (planilha) + PDF (recibos)
"""
from flask import Flask, render_template_string, request, jsonify
import io, re, unicodedata, os

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 100 * 1024 * 1024  # 100 MB

# ─────────────────────────────────────────────
# UTILITÁRIOS
# ─────────────────────────────────────────────

def norm(s: str) -> str:
    """Normaliza nome: maiúsculo, sem acento, espaços colapsados."""
    s = s.upper().strip()
    s = "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")
    return " ".join(s.split())

def brl(s) -> float:
    """Converte string monetária BR para float."""
    s = re.sub(r"[R$\s]", "", str(s))
    s = s.replace(".", "").replace(",", ".")
    try:
        return float(s)
    except Exception:
        return 0.0

def fmt_brl(v) -> str:
    if not v:
        return "-"
    try:
        return f"R$ {float(v):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception:
        return str(v)

# ─────────────────────────────────────────────
# TOLERÂNCIAS E GRUPOS DE RUBRICAS
# ─────────────────────────────────────────────

TOLERANCIA_CENTAVOS    = 0.02   # diferença até R$ 0,02 = arredondamento
TOLERANCIA_DIVERGENCIA = 0.05   # acima disso = divergência real

RUBRIC_GROUPS = {
    "COMISSAO": [
        "COMISSAO", "COMISSOES", "COMISSAO VENDA", "COMISSAO SOBRE VENDAS",
        "COMISSAO DE VENDAS", "COMISSAO MENSAL", "COMISSAO FUNCIONARIO",
        "COMISSOES DE VENDAS",
    ],
    "DSR": [
        "DSR", "D S R", "D.S.R", "DESCANSO SEMANAL REMUNERADO",
        "REPOUSO SEMANAL REMUNERADO", "DSR SOBRE COMISSAO", "DSR COMISSAO",
        "DSR S COMISSAO", "DSR S COMISSOES",
        "REFLEXO COMISSOES DSR", "REFLEXO DSR", "REFLEXO DE DSR",
        "REFLEXO COMISSAO DSR", "REFLEXO DSR COMISSOES",
    ],
    "COMISSAO_E_DSR": [
        "COMISSAO E DSR", "COMISSOES E DSR", "COMISSAO DSR", "COMISSAO + DSR",
    ],
    "VALE_TRANSPORTE": [
        "VALE TRANSPORTE", "VT", "V T", "TRANSPORTE",
        "DESCONTO VALE TRANSPORTE", "DESC VT", "D VT",
    ],
    "VALE_ALIMENTACAO": [
        "VALE ALIMENTACAO", "VA", "V A", "ALIMENTACAO",
        "TICKET ALIMENTACAO", "VALE ALIMENTACAO", "DESCONTO VALE ALIMENTACAO",
    ],
    "VALE_REFEICAO": [
        "VALE REFEICAO", "VR", "V R", "REFEICAO",
        "TICKET REFEICAO", "DESCONTO VALE REFEICAO",
    ],
    "PLANO_SAUDE": [
        "PLANO DE SAUDE", "ASSISTENCIA MEDICA", "CONVENIO MEDICO",
        "UNIMED", "AMIL", "SULAMERICA SAUDE", "PLANO SAUDE",
    ],
    "ODONTO": [
        "PLANO ODONTOLOGICO", "ODONTO", "ODONTOLOGICO",
        "ASSISTENCIA ODONTOLOGICA", "CONVENIO ODONTO",
    ],
    "ADIANTAMENTO": [
        "ADIANTAMENTO", "ADIANTAMENTO SALARIAL", "VALE SALARIAL",
    ],
}

# Índice invertido: texto normalizado → chave do grupo
_RUBRIC_INDEX: dict = {}
def _build_rubric_index():
    def _n(t):
        t = t.upper().strip()
        t = "".join(c for c in unicodedata.normalize("NFD", t) if unicodedata.category(c) != "Mn")
        t = re.sub(r"[^\w\s]", "", t)
        return " ".join(t.split())
    for group, variants in RUBRIC_GROUPS.items():
        for v in variants:
            _RUBRIC_INDEX[_n(v)] = group
_build_rubric_index()

def normalize_rubric(text: str) -> str:
    """
    Normaliza rubrica e retorna o grupo canônico (ex: 'COMISSAO_E_DSR')
    ou a rubrica normalizada se não houver grupo correspondente.
    Remove código numérico inicial (ex: '8781 SALARIO' → 'SALARIO').
    """
    t = str(text).upper().strip()
    # Remove acento
    t = "".join(c for c in unicodedata.normalize("NFD", t) if unicodedata.category(c) != "Mn")
    # Remove código numérico inicial
    t = re.sub(r"^\d+\s+", "", t)
    # Remove pontuação
    t = re.sub(r"[^\w\s]", "", t)
    # Colapsa espaços
    t = " ".join(t.split())
    return _RUBRIC_INDEX.get(t, t)

# ─────────────────────────────────────────────
# PARSER EXCEL
# ─────────────────────────────────────────────

def parse_excel(raw: bytes, filename: str) -> dict:
    """
    Lê planilha de folha de pagamento.
    Retorna: {NOME_NORM: {salario, gratificacao, ferias_13, inss, vale, plano, emprestimo, liquido}}
    """
    rows = []
    if filename.lower().endswith(".xls"):
        import xlrd
        wb = xlrd.open_workbook(file_contents=raw)
        ws = wb.sheets()[0]
        for r in range(ws.nrows):
            rows.append([ws.cell_value(r, c) for c in range(ws.ncols)])
    else:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            rows.append(list(row))

    # Detecta colunas pelos cabeçalhos
    col = {k: None for k in ["salario", "gratif", "ferias", "inss", "vale", "plano", "emprestimo", "liquido"]}
    # Colunas de apontamentos (planilha de eventos variáveis)
    apc = {}  # chave → índice de coluna
    name_col = 0  # padrão: nomes na 1ª coluna
    for row in rows[:10]:
        cells = [str(v or "").upper().strip() for v in row]
        for i, c in enumerate(cells):
            # detecta coluna de nomes
            if c in ("COLABORADORES", "NOME", "FUNCIONÁRIO", "FUNCIONARIO", "NOME DO FUNCIONÁRIO", "NOME DO FUNCIONARIO"):
                name_col = i
            if c in ("SALÁRIO", "SALARIO") and col["salario"] is None and i > 0:
                col["salario"] = i
            if ("GRATIF" in c or "ADICION" in c) and col["gratif"] is None and i > 0:
                col["gratif"] = i
            if ("FERIAS" in c or "FÉRIAS" in c or "13" in c) and col["ferias"] is None and i > 0:
                col["ferias"] = i
            if "INSS" in c and col["inss"] is None and i > 0:
                col["inss"] = i
            if "VALE" in c and col["vale"] is None and i > 0:
                col["vale"] = i
            if "PLANO" in c and col["plano"] is None and i > 0:
                col["plano"] = i
            if "EMPRESTIMO" in c and col["emprestimo"] is None and i > 0:
                col["emprestimo"] = i
            if ("LIQUIDO" in c or "LÍQUIDO" in c) and col["liquido"] is None and i > 0:
                col["liquido"] = i
            # Apontamentos
            if "ASSIDUIDADE" in c and "assiduidade" not in apc:
                apc["assiduidade"] = i
            if "PONTUALIDADE" in c and "pontualidade" not in apc:
                apc["pontualidade"] = i
            if "GRATIF" in c and "TEMPO" in c and "gratif_tempo" not in apc:
                apc["gratif_tempo"] = i
            if ("PRÊMIO" in c or "PREMIO" in c) and "premio" not in apc:
                apc["premio"] = i
            if "VALE ALIMENT" in c and "va_desconto" not in apc:
                apc["va_desconto"] = i
            if "DESCONTO FALTA" in c and "HORA" not in c and "falta" not in apc:
                apc["falta"] = i
            if "HORAS FALTA" in c and "horas_faltas" not in apc:
                apc["horas_faltas"] = i
            if "HORA EXTRA" in c and "hora_extra" not in apc:
                apc["hora_extra"] = i
            if "NOTURNO" in c and "noturno" not in apc:
                apc["noturno"] = i
            if "ADIANTAMENTO" in c and "adiantamento" not in apc:
                apc["adiantamento"] = i
            if ("FARMÁCIA" in c or "FARMACIA" in c) and "farmacia" not in apc:
                apc["farmacia"] = i

    SKIP = {"TOTAL", "NOME", "SALÁRIO", "SALARIO", "FUNCIONÁRIO", "FUNCIONARIO", "COLABORADORES", ""}
    SKIP_KW = ["LTDA", "EPP", "S/A", "CNPJ", "LISTA DE", "CENTRO MEDICO", "PAGAMENTO", "PLANILHA"]
    employees = {}

    import datetime as _dt

    for row in rows:
        first = str(row[name_col] if name_col < len(row) else "").strip()
        if not first or first.upper() in SKIP:
            continue
        if any(kw in first.upper() for kw in SKIP_KW):
            continue
        if re.match(r"^[\d\s.,\-/]+$", first):
            continue
        if len(first.split()) < 2:
            continue
        if not re.search(r"[A-Za-zÀ-ÿ]{3}", first):
            continue

        def g(k):
            ci = col[k]
            if ci is None or ci >= len(row):
                return 0.0
            v = row[ci]
            return float(v) if isinstance(v, (int, float)) else brl(v)

        def ga(k):
            """Lê coluna de apontamento; retorna None se vazio."""
            ci = apc.get(k)
            if ci is None or ci >= len(row):
                return None
            v = row[ci]
            if v is None:
                return None
            if isinstance(v, _dt.datetime):
                return v.strftime("%d/%m/%Y")
            if isinstance(v, _dt.timedelta):
                ts = int(v.total_seconds())
                return f"{ts//3600}:{(ts%3600)//60:02d}"
            try:
                f = float(v)
                return f if f != 0 else None
            except Exception:
                s = str(v).strip()
                return s if s else None

        # fallback posicional — inclui zeros para manter posições corretas
        nums = [float(v) for v in row[name_col+1:] if isinstance(v, (int, float))]

        aponts = {k: ga(k) for k in apc if ga(k) is not None}

        employees[norm(first)] = {
            "salario":      g("salario")     or (nums[0] if nums else 0),
            "gratificacao": g("gratif")      or (nums[1] if len(nums) > 1 else 0),
            "ferias_13":    g("ferias")      or (nums[2] if len(nums) > 2 else 0),
            "inss":         g("inss"),
            "vale":         g("vale"),
            "plano":        g("plano"),
            "emprestimo":   g("emprestimo"),
            "liquido":      g("liquido")     or (nums[-1] if nums else 0),
            "has_liquido":  col["liquido"] is not None,
            "apontamentos": aponts,
        }

    return employees

# ─────────────────────────────────────────────
# PARSER PDF (Recibos)
# ─────────────────────────────────────────────

def parse_pdf(raw: bytes) -> dict:
    """
    Lê PDF de recibos de pagamento.
    Retorna: {NOME_NORM: {liquido, total_vencimentos, total_descontos, verbas, tipo, has_gratif}}
    """
    import pdfplumber

    employees = {}
    seen = set()

    with pdfplumber.open(io.BytesIO(raw)) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""

            # ── Extrai nome do funcionário ──────────────────────────────────
            # Padrão: código numérico + NOME EM CAPS + CBO (6 dígitos)
            # Exemplo: "58 ANDREIA PEREIRA BARBOSA 514320 1 1"
            nm = re.search(
                r"\b\d{1,3}\s+([A-Z][A-Z ]{4,50}?)\s+\d{6}\b",
                text,
            )

            if not nm:
                continue

            emp_name = nm.group(1).strip()
            emp_norm = norm(emp_name)

            if emp_norm in seen or len(emp_norm.split()) < 2:
                continue
            seen.add(emp_norm)

            # ── Tipo de recibo ──────────────────────────────────────────────
            tipo = "mensal"
            if re.search(r"13[oOº°].*adiantamento|adiantamento.*13", text[:600], re.I):
                tipo = "13_adiantamento"
            elif re.search(r"f[eé]rias", text[:600], re.I):
                tipo = "ferias"

            # ── Totais ──────────────────────────────────────────────────────
            # "Valor Líquido 1.143,17"  (o \uf0f0 é a seta Wingdings do sistema)
            liq_m = re.search(
                r"Valor\s+L[íi]quido[^0-9]*([\d.,]+)",
                text, re.I,
            )
            # "Total de Descontos\n883,08" — aparece ANTES do líquido no texto
            td_m = re.search(
                r"Total\s+de\s+Descontos[^0-9]*([\d.,]+)",
                text, re.I,
            )

            liquido    = brl(liq_m.group(1)) if liq_m else 0.0
            total_desc = brl(td_m.group(1)) if td_m else 0.0
            total_venc = round(liquido + total_desc, 2)   # sempre verdadeiro contabilmente

            # ── Verbas ──────────────────────────────────────────────────────
            verbas = []
            # Padrão: "8781 DIAS NORMAIS 30,00 1.621,00"
            # Referência aceita ponto (ex: "1.200,00" para Unimed/empréstimos)
            for m in re.finditer(
                r"(\d{2,4})\s+([A-ZÀ-Ü][A-ZÀ-Ü .%º/°()]+?)\s+([\d:.,]+)\s+([\d.,]+)",
                text,
            ):
                desc = m.group(2).strip()
                ref  = m.group(3)
                # Referência legítima: tem vírgula (valor monetário) OU dois pontos (horas, ex: "3:41")
                if "," not in ref and ":" not in ref:
                    continue
                # Ignora palavras-chave que não são verbas
                if any(kw in desc for kw in ["CNPJ", "BASE CALC", "SAL. CONTR", "F.G.T.S"]):
                    continue
                valor = brl(m.group(4))
                # Valores < 1 são ruídos (ex: filial = "1")
                if valor < 1:
                    continue
                verbas.append({
                    "codigo": m.group(1),
                    "descricao": desc,
                    "referencia": ref,
                    "valor": valor,
                })

            # Captura verbas sem código (ex: "DESCONTO VALE ALIMENTAÇÃO 18,00 18,00")
            _SKIP_CODELESS = {"CODIGO", "DESCRIÇÃO", "DESCRICAO", "TOTAL", "REFERENCIA", "REFERÊNCIA",
                              "VALOR LIQUIDO", "VALOR LÍQUIDO", "BASE CALC", "SAL.", "SALARIO", "SALÁRIO"}
            for m in re.finditer(
                r"^([A-ZÀ-Ü][A-ZÀ-Ü ]{3,50}?)\s+([\d:.,]+)\s+([\d.,]+)\s*$",
                text, re.MULTILINE
            ):
                desc_c = m.group(1).strip()
                ref_c  = m.group(2)
                if "," not in ref_c and ":" not in ref_c:
                    continue
                if len(desc_c.split()) < 2:
                    continue
                if any(kw in desc_c.upper() for kw in _SKIP_CODELESS):
                    continue
                if any(kw in desc_c.upper() for kw in ["CNPJ", "F.G.T.S", "BASE"]):
                    continue
                # Evita duplicar verbas já capturadas pelo padrão com código
                if any(v["descricao"].upper() == desc_c.upper() for v in verbas):
                    continue
                valor_c = brl(m.group(3))
                if valor_c < 1:
                    continue
                verbas.append({
                    "codigo": "",
                    "descricao": desc_c,
                    "referencia": ref_c,
                    "valor": valor_c,
                })

            has_gratif = any(
                re.search(r"GRATIF|PREMIACAO|PREMIO", v["descricao"].upper())
                for v in verbas
            )
            gratif_valor = sum(
                v["valor"] for v in verbas
                if re.search(r"GRATIF|PREMIACAO|PREMIO", v["descricao"].upper())
            )

            employees[emp_norm] = {
                "nome_original":    emp_name.title(),
                "tipo":             tipo,
                "liquido":          liquido,
                "total_vencimentos": total_venc,
                "total_descontos":  total_desc,
                "has_gratif":       has_gratif,
                "gratif_valor":     gratif_valor,
                "verbas":           verbas,
            }

    return employees

# ─────────────────────────────────────────────
# PARSER WORD
# ─────────────────────────────────────────────

def parse_word(raw: bytes) -> dict:
    """
    Lê documento Word com instruções da folha.
    Retorna: {gratificacoes, descontos, obs, decimo_terceiro}
    """
    from docx import Document

    doc = Document(io.BytesIO(raw))
    text = "\n".join(p.text for p in doc.paragraphs)

    result = {"gratificacoes": {}, "descontos": {}, "obs": [], "decimo_terceiro": [], "comissao_e_dsr": {}}

    # Gratificações
    m = re.search(
        r"Gratifica[çc][õo]es?:?(.*?)(?:Funcionar|Férias|Descontos|Goiânia|$)",
        text, re.I | re.S,
    )
    if m:
        for entry in re.finditer(
            r"([A-Za-zÀ-ÿ][A-Za-zÀ-ÿ\s]{3,40}?)\s*[\-–\(]?\s*\(?([\d.,]{3,})\)?",
            m.group(1),
        ):
            name, val = entry.group(1).strip(), brl(entry.group(2))
            if val > 0 and len(name.split()) >= 2:
                result["gratificacoes"][norm(name)] = val

    # Descontos (Unimed, etc.)
    m2 = re.search(r"Descontos?:?(.*?)(?:Goiânia|Ass|$)", text, re.I | re.S)
    if m2:
        tipo_atual = "Desconto"
        for line in m2.group(1).split("\n"):
            lt = line.strip()
            if not lt:
                continue
            if re.match(r"^([A-Za-zÀ-ÿ]+):?\s*$", lt):
                tipo_atual = lt.rstrip(":")
                continue
            entry = re.search(
                r"([A-Za-zÀ-ÿ][A-Za-zÀ-ÿ\s]{4,40}?)\s*[\-–]\s*\(?([\d.,]{3,})\)?", lt
            )
            if entry:
                name, val = entry.group(1).strip(), brl(entry.group(2))
                if val > 0 and len(name.split()) >= 2:
                    nn = norm(name)
                    result["descontos"].setdefault(nn, {})[tipo_atual] = val

    # Adiantamento 13º
    result["decimo_terceiro"] = [
        norm(x)
        for x in re.findall(
            r"(?:13[oOº°]|décimo\s+terceiro)[^\w]*([A-Za-zÀ-ÿ][A-Za-zÀ-ÿ\s]{5,40})",
            text, re.I,
        )
        if len(x.split()) >= 2
    ]

    # Observações
    result["obs"] = [
        l.strip() for l in text.split("\n") if re.match(r"OBS|Obs", l.strip())
    ]

    # Fallback: formato lista simples (título na 1ª linha, depois nome\nvalor alternados)
    # Ex: "comissões e DSR\nNadyane\n1592,11\nJuliana\n1255,49"
    if not result["gratificacoes"] and not result["descontos"]:
        lines = [l.strip() for l in text.split("\n") if l.strip()]
        if len(lines) >= 3:
            # Detecta o tipo da seção pelo título (linha 0)
            titulo_norm = normalize_rubric(lines[0])
            is_comissao_dsr = titulo_norm in ("COMISSAO_E_DSR", "COMISSAO E DSR")

            # Detecta se é padrão nome/valor: linhas alternadas texto/número
            pares = []
            i = 1  # pula título (linha 0)
            while i < len(lines) - 1:
                nome_linha = lines[i]
                val_linha = lines[i + 1]
                val_clean = val_linha.replace(".", "").replace(",", ".").replace(" ", "")
                if re.match(r"^\d+(\.\d+)?$", val_clean):
                    pares.append((nome_linha, val_linha))
                    i += 2
                else:
                    i += 1
            if pares:
                for nome, val_str in pares:
                    val = brl(val_str)
                    name = norm(nome)
                    if val > 0 and len(name) >= 2:  # aceita nome com 1 palavra
                        if is_comissao_dsr:
                            result["comissao_e_dsr"][name] = val
                        else:
                            result["gratificacoes"][name] = val

    return result

# ─────────────────────────────────────────────
# MOTOR DE COMPARAÇÃO
# ─────────────────────────────────────────────

def match_names(excel: dict, pdf: dict) -> dict:
    """
    Casa nomes do Excel (abreviados) com nomes do PDF (completos).
    Retorna: {excel_name: pdf_name}
    """
    mapping = {}
    used_pdf = set()

    for en in excel:
        en_words = en.split()
        best = None
        # Tenta prefixo exato: "ANDREIA PEREIRA" → "ANDREIA PEREIRA BARBOSA"
        for pn in pdf:
            if pn in used_pdf:
                continue
            pn_words = pn.split()
            if pn_words[:len(en_words)] == en_words:
                best = pn
                break
        # Fallback: verifica se as 2 primeiras palavras batem
        if not best:
            for pn in pdf:
                if pn in used_pdf:
                    continue
                pn_w = pn.split()
                en_w = en.split()
                if pn_w[:2] == en_w[:2]:
                    best = pn
                    break
        if best:
            mapping[en] = best
            used_pdf.add(best)

    return mapping


def compare(excel: dict, pdf: dict, word: dict) -> dict:
    # Casa nomes Excel (abreviados) ↔ PDF (completos)
    name_map = match_names(excel, pdf)          # excel_name → pdf_name
    rev_map  = {v: k for k, v in name_map.items()}  # pdf_name → excel_name

    # Todos os funcionários (usando nomes do Excel como chave canônica quando possível)
    canonical = {}  # canonical_name → {excel_key, pdf_key}
    for en in excel:
        pn = name_map.get(en)
        canonical[en] = {"excel_key": en, "pdf_key": pn}
    for pn in pdf:
        en = rev_map.get(pn)
        if not en:
            canonical[pn] = {"excel_key": None, "pdf_key": pn}

    all_names = sorted(canonical)

    report = {
        "resumo": {"total": len(all_names), "divergencias": 0, "ok": 0},
        "funcionarios": [],
        "observacoes": (word or {}).get("obs", []),
        "word_gratificacoes": {k: fmt_brl(v) for k, v in (word or {}).get("gratificacoes", {}).items()},
        "word_descontos": (word or {}).get("descontos", {}),
    }

    for name in all_names:
        keys    = canonical[name]
        exc_key = keys["excel_key"]
        pdf_key = keys["pdf_key"]
        exc = excel.get(exc_key) if exc_key else None
        rec = pdf.get(pdf_key)   if pdf_key else None

        nome_exibir = (rec or {}).get("nome_original") or (exc_key or name).title()

        emp = {
            "nome":        name,
            "nome_exibir": nome_exibir,
            "status":      "OK",
            "divs":        [],
            "dados_excel": exc,
            "dados_recibo": rec,
        }

        # ── Presença ────────────────────────────────────────────────────────
        # Só aponta ausência se o respectivo tipo de arquivo foi enviado
        if not exc and excel:
            emp["divs"].append({
                "g": "alta",
                "tipo": "Ausente na planilha",
                "desc": "Funcionário tem recibo mas não está na planilha Excel.",
            })
        if not rec and pdf:
            emp["divs"].append({
                "g": "alta",
                "tipo": "Sem recibo PDF",
                "desc": "Funcionário está na planilha mas não há recibo PDF.",
            })

        # ── Valor líquido ───────────────────────────────────────────────────
        if exc and rec and exc.get("has_liquido", True):
            el = exc.get("liquido", 0)
            rl = rec.get("liquido", 0)
            if el > 0 and abs(el - rl) > TOLERANCIA_DIVERGENCIA:
                emp["divs"].append({
                    "g": "alta",
                    "tipo": "Líquido divergente",
                    "desc": (
                        f"Planilha: {fmt_brl(el)} | Recibo: {fmt_brl(rl)} "
                        f"| Diferença: {fmt_brl(abs(el - rl))}"
                    ),
                })

        # ── Apontamentos vs Verbas do recibo ────────────────────────────────
        if exc and rec:
            aponts = exc.get("apontamentos", {})
            verbas = rec.get("verbas", [])

            def _find_verba(keywords, codigo=None):
                # Busca por código + keyword (evita falsos positivos de parsing)
                if codigo:
                    for v in verbas:
                        if (v.get("codigo") == str(codigo) and
                                any(kw.upper() in v.get("descricao", "").upper() for kw in keywords)):
                            return v
                # Fallback: keyword apenas
                for v in verbas:
                    if any(kw.upper() in v.get("descricao", "").upper() for kw in keywords):
                        return v
                return None

            def _check_bonus(chave, label, keywords, codigo=None):
                val = aponts.get(chave)
                if not val:
                    return
                try:
                    val = float(val)
                except Exception:
                    return
                v = _find_verba(keywords, codigo)
                if not v:
                    emp["divs"].append({
                        "g": "alta",
                        "tipo": f"{label} ausente no recibo",
                        "desc": f"Planilha indica {label} de {fmt_brl(val)}, mas não encontrado no recibo.",
                    })
                elif abs(v.get("valor", 0) - val) > TOLERANCIA_DIVERGENCIA:
                    emp["divs"].append({
                        "g": "media",
                        "tipo": f"{label} divergente",
                        "desc": f"Planilha: {fmt_brl(val)} | Recibo: {fmt_brl(v.get('valor',0))}",
                    })

            _check_bonus("pontualidade", "Pontualidade", ["PONTUALIDADE"], "221")
            _check_bonus("assiduidade",  "Assiduidade",  ["ASSIDUIDADE"],  "222")
            _check_bonus("gratif_tempo", "Gratificação Tempo de Serviço", ["GRATIF", "TEMPO"], "228")
            _check_bonus("premio",       "Prêmio",       ["PREMIO", "PREMIA"], None)
            _check_bonus("va_desconto",  "Vale Alimentação", ["VALE ALIMENT"], "204")
            _check_bonus("adiantamento", "Adiantamento Salarial", ["ADIANT"], None)
            _check_bonus("farmacia",     "Farmácia", ["FARMACIA", "FARMÁCIA"], None)

            # Falta: só verifica presença da verba, não o valor (que é data na planilha)
            if aponts.get("falta"):
                v = _find_verba(["FALTA", "DIAS FALTA", "DESCONTO FALTA"], "8792")
                if not v:
                    emp["divs"].append({
                        "g": "media",
                        "tipo": "Desconto de falta ausente no recibo",
                        "desc": f"Planilha indica falta em {aponts['falta']}, mas não há desconto de falta no recibo.",
                    })

            # Horas faltas: verifica presença
            if aponts.get("horas_faltas"):
                v = _find_verba(["HORA FALTA", "HORAS FALTA", "FALTAS PARC"], "8069")
                if not v:
                    emp["divs"].append({
                        "g": "media",
                        "tipo": "Horas de falta parcial ausente no recibo",
                        "desc": f"Planilha indica {aponts['horas_faltas']}h de falta parcial, mas não encontrado no recibo.",
                    })

            # Hora extra: verifica presença
            if aponts.get("hora_extra"):
                v = _find_verba(["HORA EXTRA", "H.EXTRA", "HORAS EXTRA"])
                if not v:
                    emp["divs"].append({
                        "g": "media",
                        "tipo": "Hora extra ausente no recibo",
                        "desc": f"Planilha indica {aponts['hora_extra']}h extra, mas não encontrado no recibo.",
                    })

            # Adicional noturno: verifica presença
            if aponts.get("noturno"):
                v = _find_verba(["NOTURNO", "ADICIONAL NOT"])
                if not v:
                    emp["divs"].append({
                        "g": "media",
                        "tipo": "Adicional noturno ausente no recibo",
                        "desc": f"Planilha indica {aponts['noturno']}h noturno, mas não encontrado no recibo.",
                    })

        # ── Gratificações (Word → Recibo) ────────────────────────────────────
        if word:
            # Word pode usar nome abreviado, completo ou apenas primeiro nome
            wg = word.get("gratificacoes", {})
            gratif = wg.get(exc_key or name) or wg.get(pdf_key or name) or 0
            # Fallback: match por primeiro nome (quando Word usa só "NADYANE")
            if not gratif:
                first_name = (exc_key or pdf_key or name or "").split()[0] if (exc_key or pdf_key or name) else ""
                if first_name:
                    for wk, wv in wg.items():
                        if wk.split()[0] == first_name:
                            gratif = wv
                            break
            if gratif > 0:
                if rec and not rec.get("has_gratif", False):
                    emp["divs"].append({
                        "g": "alta",
                        "tipo": "Gratificação ausente no recibo",
                        "desc": (
                            f"Gratificação de {fmt_brl(gratif)} consta no Word "
                            f"mas NÃO aparece no recibo como verba separada."
                        ),
                    })
                exc_g = (exc or {}).get("gratificacao", 0)
                if exc_g > 0 and abs(exc_g - gratif) > TOLERANCIA_DIVERGENCIA:
                    emp["divs"].append({
                        "g": "media",
                        "tipo": "Valor de gratificação divergente",
                        "desc": f"Word: {fmt_brl(gratif)} | Planilha: {fmt_brl(exc_g)}",
                    })

            # ── Descontos especiais (Word → Recibo) ─────────────────────────
            wd = word.get("descontos", {})
            word_descs = wd.get(exc_key or name) or wd.get(pdf_key or name) or {}
            for tipo_desc, val in word_descs.items():
                verbas = (rec or {}).get("verbas", [])
                found = any(
                    tipo_desc.upper() in v.get("descricao", "").upper()
                    or abs(v.get("valor", 0) - val) < TOLERANCIA_DIVERGENCIA
                    for v in verbas
                )
                if not found and rec:
                    emp["divs"].append({
                        "g": "media",
                        "tipo": f"Desconto '{tipo_desc}' não localizado",
                        "desc": (
                            f"Desconto {tipo_desc} de {fmt_brl(val)} mencionado no Word "
                            f"não foi encontrado no recibo."
                        ),
                    })

        # ── Comissão + DSR (Word → Recibo) ──────────────────────────────────
        if word:
            cdsr = word.get("comissao_e_dsr", {})
            val_word = cdsr.get(exc_key or name) or cdsr.get(pdf_key or name) or 0
            if not val_word:
                first = (exc_key or pdf_key or name or "").split()[0] if (exc_key or pdf_key or name) else ""
                if first:
                    for wk, wv in cdsr.items():
                        if wk.split()[0] == first:
                            val_word = wv
                            break
            if val_word and val_word > 0 and rec:
                val_recibo = sum(
                    v["valor"] for v in rec.get("verbas", [])
                    if normalize_rubric(v["descricao"]) in ("COMISSAO", "DSR", "COMISSAO_E_DSR")
                )
                diff = round(abs(val_word - val_recibo), 2)
                mem = {
                    "word_valor": val_word,
                    "recibo_verbas": [
                        {"desc": v["descricao"], "valor": v["valor"]}
                        for v in rec.get("verbas", [])
                        if normalize_rubric(v["descricao"]) in ("COMISSAO", "DSR", "COMISSAO_E_DSR")
                    ],
                    "recibo_total": val_recibo,
                    "diferenca": diff,
                }
                if diff <= TOLERANCIA_CENTAVOS:
                    emp["memoria_comissao_dsr"] = {**mem, "status": "OK"}
                elif val_word > val_recibo:
                    emp["divs"].append({
                        "g": "alta",
                        "tipo": "DIFERENÇA A PAGAR",
                        "desc": (
                            f"Comissão+DSR — Word: {fmt_brl(val_word)} | "
                            f"Recibo: {fmt_brl(val_recibo)} | Falta: {fmt_brl(diff)}"
                        ),
                        "memoria": mem,
                    })
                else:
                    emp["divs"].append({
                        "g": "media",
                        "tipo": "PAGO A MAIOR",
                        "desc": (
                            f"Comissão+DSR — Word: {fmt_brl(val_word)} | "
                            f"Recibo: {fmt_brl(val_recibo)} | Excesso: {fmt_brl(diff)}"
                        ),
                        "memoria": mem,
                    })

        if emp["divs"]:
            emp["status"] = "DIVERGENTE"
            report["resumo"]["divergencias"] += 1
        else:
            report["resumo"]["ok"] += 1

        report["funcionarios"].append(emp)

    # ── Possíveis homônimos ──────────────────────────────────────────────
    possiveis = []
    nomes = [c for c in all_names if canonical[c]["pdf_key"]]
    for i in range(len(nomes)):
        for j in range(i + 1, len(nomes)):
            a_words = nomes[i].split()
            b_words = nomes[j].split()
            if (len(a_words) >= 2 and len(b_words) >= 2
                    and a_words[:2] == b_words[:2] and nomes[i] != nomes[j]):
                possiveis.append({
                    "nomes": [nomes[i], nomes[j]],
                    "aviso": "Possível homônimo: mesmas 2 primeiras palavras",
                })
    report["possiveis_homonimos"] = possiveis

    return report

# ─────────────────────────────────────────────
# PARSER FATURA PLANO DE SAÚDE (Unimed analítico)
# ─────────────────────────────────────────────

def fix_spaced(text: str) -> str:
    """Remove espaços entre dígitos (formato Unimed: '6 2 9 , 3 5' → '629,35')."""
    for _ in range(15):
        text = re.sub(r"(\d) (\d)", r"\1\2", text)
    text = re.sub(r"(\d)\s*,\s*(\d)", r"\1,\2", text)
    text = re.sub(r"(\d)\s+\.\s*(\d)", r"\1.\2", text)
    text = re.sub(r"(\d)\s*\.\s+(\d)", r"\1.\2", text)
    return text


def _page_lines_smart(page) -> list:
    """
    Extrai linhas de uma página usando posições de caracteres para reconstruir
    limites de palavras reais. Funciona com PDFs onde cada letra é espaçada
    individualmente (ex: Unimed analítico), detectando que word-gaps têm ~2x
    o espaçamento de letter-gaps.
    """
    from collections import defaultdict
    try:
        from statistics import median
    except ImportError:
        def median(lst):
            s = sorted(lst)
            n = len(s)
            return s[n // 2] if n % 2 else (s[n // 2 - 1] + s[n // 2]) / 2

    chars = page.chars
    if not chars:
        return []

    rows = defaultdict(list)
    for c in chars:
        y_key = round(c["doctop"] / 2) * 2
        rows[y_key].append(c)

    lines = []
    for y in sorted(rows.keys()):
        row = [c for c in rows[y] if c["text"].strip()]
        if not row:
            continue
        row.sort(key=lambda c: c["x0"])
        if len(row) == 1:
            lines.append(row[0]["text"])
            continue

        gaps = [row[i + 1]["x0"] - row[i]["x0"] for i in range(len(row) - 1)]
        med = median(gaps)
        threshold = max(med * 1.6, 1.5)

        text = row[0]["text"]
        for i, gap in enumerate(gaps):
            if gap > threshold:
                text += " "
            text += row[i + 1]["text"]
        lines.append(text)

    return lines


def _merge_fatura(base: dict, new: dict) -> dict:
    """Acumula valores da fatura — soma ao invés de sobrescrever (multi-arquivo/mês)."""
    for key, val in new.items():
        if key in base:
            base[key]["mensalidade"] = round(base[key]["mensalidade"] + val["mensalidade"], 2)
            base[key]["mensalidade_dependentes"] = round(
                base[key].get("mensalidade_dependentes", 0) + val.get("mensalidade_dependentes", 0), 2
            )
            base[key]["sos_tam"] = round(base[key].get("sos_tam", 0) + val.get("sos_tam", 0), 2)
            base[key]["total"] = round(base[key]["total"] + val["total"], 2)
            dep = base[key].get("dependentes")
            new_dep = val.get("dependentes", [])
            if isinstance(dep, list):
                base[key]["dependentes"] = dep + new_dep
            else:
                base[key]["dependentes"] = (dep or 0) + len(new_dep)
        else:
            base[key] = dict(val)
    return base


def _merge_extrato(base: dict, new: dict) -> dict:
    """Acumula descontos do extrato — soma ao invés de sobrescrever (multi-arquivo/mês)."""
    for key, val in new.items():
        if key in base:
            base[key]["plano_descontado"] = round(
                base[key]["plano_descontado"] + val["plano_descontado"], 2
            )
            if val.get("salario", 0) > base[key].get("salario", 0):
                base[key]["salario"] = val["salario"]
        else:
            base[key] = dict(val)
    return base


def parse_plano_fatura(raw: bytes, filtro_linha: str = "MENSALIDADE") -> dict:
    """
    Lê fatura/relatório de benefício e extrai valor por titular.
    Suporta formato Unimed analítico (letras espaçadas individualmente) e
    documentos genéricos com código-nome-valor.
    """
    import pdfplumber

    filtro = filtro_linha.upper().strip()
    titulares = {}
    titular_atual = None

    with pdfplumber.open(io.BytesIO(raw)) as pdf:
        for page in pdf.pages:
            # Usa extração inteligente por posição de chars para reconstruir
            # word boundaries reais (resolve formato Unimed tudo-espaçado)
            lines = _page_lines_smart(page)
            for line in lines:
                line = line.strip()
                if not line:
                    continue

                filtro_ok  = filtro in line.upper()
                sos_tam_ok = bool(re.search(r"\bSOS\b|\bTAM\b", line.upper()))
                if not filtro_ok and not sos_tam_ok:
                    continue

                # Normaliza espaços em dígitos que ainda estejam separados
                line_fixed = fix_spaced(line)

                valores = re.findall(r"(?<!\d)[\d.]+,\d{2}(?!\d)", line_fixed)
                valores_validos = [v for v in valores if brl(v) >= 1.0]
                total_val = brl(valores_validos[-1]) if valores_validos else 0.0

                if not filtro_ok and sos_tam_ok:
                    if titular_atual and titular_atual in titulares and total_val > 0:
                        titulares[titular_atual]["sos_tam"] += total_val
                    continue

                if total_val == 0:
                    continue

                cod_m = re.search(r"(\d{4}\.\d{4}\.\d{6})-(\d{2,3})", line_fixed)

                if cod_m:
                    codigo_base = cod_m.group(1)
                    sufixo      = cod_m.group(2)

                    pos_cod  = line_fixed.find(cod_m.group(0))
                    pos_filt = line_fixed.upper().find(filtro.split()[0])
                    if pos_filt < 0:
                        pos_filt = len(line_fixed)
                    trecho = line_fixed[pos_cod + len(cod_m.group(0)):pos_filt].strip()

                    nome_raw = re.sub(r"^\s*[AIER]\s+", "", trecho).strip()
                    nome_raw = re.sub(r"\s+[AIER]\s*$", "", nome_raw).strip()
                    nome_raw = re.sub(r"\b\d+\b", "", nome_raw).strip()
                    nome_raw = re.sub(r"\s{2,}", " ", nome_raw).strip()

                    if sufixo == "00":
                        titular_atual = codigo_base
                        titulares[codigo_base] = {
                            "nome_original": nome_raw.title() if nome_raw else codigo_base,
                            "nome_norm":     norm(nome_raw) if nome_raw else codigo_base,
                            "mensalidade":   total_val,
                            "sos_tam":       0.0,
                            "dependentes":   [],
                        }
                    else:
                        if titular_atual and titular_atual in titulares:
                            titulares[titular_atual]["dependentes"].append({
                                "nome": nome_raw.title(),
                                "valor": total_val,
                            })
                else:
                    pos_filt = line_fixed.upper().find(filtro.split()[0])
                    nome_raw = line_fixed[:pos_filt].strip() if pos_filt > 0 else ""
                    nome_raw = re.sub(r"\b\d+\b", "", nome_raw).strip()
                    nome_raw = re.sub(r"^\s*[AIER]\s+", "", nome_raw).strip()
                    nome_raw = re.sub(r"\s+[AIER]\s*$", "", nome_raw).strip()
                    nome_raw = re.sub(r"\s{2,}", " ", nome_raw).strip()

                    if nome_raw and len(nome_raw.split()) >= 2:
                        nn = norm(nome_raw)
                        titular_atual = nn
                        titulares[nn] = {
                            "nome_original": nome_raw.title(),
                            "nome_norm":     nn,
                            "mensalidade":   total_val,
                            "sos_tam":       0.0,
                            "dependentes":   [],
                        }
                    elif titular_atual and titular_atual in titulares:
                        titulares[titular_atual]["dependentes"].append({
                            "nome": "",
                            "valor": total_val,
                        })

    result = {}
    for dados in titulares.values():
        nn = dados["nome_norm"]
        if not nn:
            continue
        total_dep = sum(d["valor"] for d in dados["dependentes"])
        result[nn] = {
            "nome_original":          dados["nome_original"],
            "mensalidade":            dados["mensalidade"],
            "sos_tam":                dados["sos_tam"],
            "total":                  round(dados["mensalidade"] + total_dep + dados["sos_tam"], 2),
            "mensalidade_dependentes": total_dep,
            "dependentes":            dados["dependentes"],
        }

    return result


# ─────────────────────────────────────────────
# PARSER EXTRATO DE FOLHA (desconto 8111)
# ─────────────────────────────────────────────

def parse_extrato_plano(raw: bytes, codigo: str = "8111") -> dict:
    """
    Lê extrato de folha e extrai o evento pelo código informado.
    Retorna: {NOME_NORM: {nome_original, plano_descontado, salario}}
    """
    import pdfplumber

    result = {}
    cod = str(codigo).strip()

    with pdfplumber.open(io.BytesIO(raw)) as pdf:
        full_text = ""
        for page in pdf.pages:
            full_text += (page.extract_text() or "") + "\n"

    blocos = re.split(r"(?=Empr\.:?\s*\d+[A-Z])", full_text)

    for bloco in blocos:
        nm = re.search(r"Empr\.:?\s*\d+([A-Z][A-Z\s]+?)(?:\s+Situa[çc][aã]o\s*:|\s+CPF\s*:)", bloco)
        if not nm:
            continue

        nome = nm.group(1).strip()
        nome = re.sub(r"\s{2,}", " ", nome)

        sal_m = re.search(r"Sal[aá]rio:\s*([\d.,]+)", bloco, re.I)
        salario = brl(sal_m.group(1)) if sal_m else 0.0

        # Busca TODAS as ocorrências do código e soma (desconto pode aparecer mais de uma vez)
        matches = re.findall(
            rf"{re.escape(cod)}\s+[A-Z][A-Z0-9\s./ºÇÃÕÁÉÍÓÚ]*?\s+([\d.,]+)\s+([\d.,]+)\s*D?",
            bloco, re.I
        )
        valor_descontado = round(sum(brl(m[1]) for m in matches), 2)

        if not matches and salario == 0.0:
            continue

        chave = norm(nome)
        if chave in result:
            # Mesmo funcionário em outro mês do mesmo arquivo — acumula
            result[chave]["plano_descontado"] = round(
                result[chave]["plano_descontado"] + valor_descontado, 2
            )
            if salario > result[chave].get("salario", 0):
                result[chave]["salario"] = salario
        else:
            result[chave] = {
                "nome_original": nome.title(),
                "plano_descontado": valor_descontado,
                "salario": salario,
            }

    return result


def parse_referencia_simples(raw: bytes, filename: str) -> dict:
    """
    Lê documento de referência em formato Excel simples (nome | valor).
    Retorna: {NOME_NORM: {nome_original, mensalidade, total}}
    """
    rows = []
    if filename.lower().endswith(".xls"):
        import xlrd
        wb = xlrd.open_workbook(file_contents=raw)
        ws = wb.sheets()[0]
        for r in range(ws.nrows):
            rows.append([ws.cell_value(r, c) for c in range(ws.ncols)])
    else:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            rows.append(list(row))

    result = {}
    for row in rows:
        if not row or not row[0]:
            continue
        nome_raw = str(row[0]).strip()
        if len(nome_raw) < 3 or not re.search(r"[A-Za-zÀ-ÿ]{2}", nome_raw):
            continue
        # Procura o primeiro valor numérico na linha
        valor = 0.0
        for cell in row[1:]:
            if isinstance(cell, (int, float)) and cell > 0:
                valor = float(cell)
                break
            elif isinstance(cell, str):
                v = brl(cell)
                if v > 0:
                    valor = v
                    break
        if valor > 0:
            nn = norm(nome_raw)
            result[nn] = {
                "nome_original": nome_raw.title(),
                "mensalidade": valor,
                "mensalidade_dependentes": 0.0,
                "sos_tam": 0.0,
                "total": valor,
                "dependentes": 0,
            }
    return result


# ─────────────────────────────────────────────
# COMPARAÇÃO PLANO DE SAÚDE
# ─────────────────────────────────────────────

def _abbrev_match(fw_list: list, ew_list: list) -> bool:
    """
    Verifica se fw_list (nomes da fatura, possivelmente abreviados) representa
    a mesma pessoa que ew_list (nomes completos do extrato).
    Regras:
    - Primeira palavra deve coincidir exatamente
    - Palavras de 1 letra são tratadas como inicial — "O" bate com "OLIVEIRA"
    - Palavras do extrato podem ser puladas (nomes do meio ausentes na fatura)
    """
    if not fw_list or not ew_list or fw_list[0] != ew_list[0]:
        return False
    fi, ei = 1, 1
    while fi < len(fw_list) and ei < len(ew_list):
        fw, ew = fw_list[fi], ew_list[ei]
        if fw == ew:
            fi += 1; ei += 1
        elif len(fw) == 1 and ew.startswith(fw):
            fi += 1; ei += 1
        else:
            ei += 1  # pula palavra do extrato (nome do meio não abreviado)
    return fi == len(fw_list)


def match_names_beneficio(fatura: dict, extrato: dict) -> dict:
    """
    Casa nomes da fatura (truncados/abreviados) com nomes completos do extrato.
    Retorna: {fatura_key → extrato_key}
    """
    mapping = {}
    used = set()

    for fk in fatura:
        fk_words = fk.split()
        best = None

        # 1) Prefixo exato (fatura é prefixo do extrato)
        for ek in extrato:
            if ek in used: continue
            ek_words = ek.split()
            if ek_words[:len(fk_words)] == fk_words:
                best = ek; break

        # 2) Match com abreviações (ex: "christiane o santos" = "christiane oliveira dos santos")
        if not best:
            for ek in extrato:
                if ek in used: continue
                if _abbrev_match(fk_words, ek.split()):
                    best = ek; break

        # 3) Primeiras 2 palavras exatas (fallback)
        if not best:
            for ek in extrato:
                if ek in used: continue
                if ek.split()[:2] == fk_words[:2]:
                    best = ek; break

        if best:
            mapping[fk] = best
            used.add(best)

    return mapping


def compare_plano_saude(fatura: dict, extrato: dict, regra: dict = None) -> dict:
    """
    Compara valor esperado (fatura ou regra) com o que foi descontado no extrato.
    regra: {"tipo": "fatura"|"pct_fatura"|"pct_salario"|"fixo", "valor": float}
    """
    regra = regra or {"tipo": "fatura", "valor": 0.0}
    name_map = match_names_beneficio(fatura, extrato)
    rev_map  = {v: k for k, v in name_map.items()}

    todos = set(fatura.keys()) | {rev_map.get(ek, ek) for ek in extrato}

    resultados = []
    total_esperado = 0.0
    total_extrato  = 0.0
    divergentes    = 0

    def calc_esperado(fat, ext):
        """Calcula valor esperado baseado na regra."""
        tipo = regra["tipo"]
        pct  = regra["valor"] / 100.0
        fixo = regra["valor"]
        if tipo == "fatura":
            return (fat or {}).get("total", 0.0)
        elif tipo == "pct_fatura":
            return round((fat or {}).get("total", 0.0) * pct, 2)
        elif tipo == "pct_salario":
            return round((ext or {}).get("salario", 0.0) * pct, 2)
        elif tipo == "fixo":
            return fixo if (ext or fat) else 0.0
        return 0.0

    processed_ek = set()  # rastreia extrato keys já processadas

    for fk in sorted(todos):
        ek  = name_map.get(fk)
        fat = fatura.get(fk)
        ext = extrato.get(ek) if ek else None
        if not ext and fk in extrato:
            ext = extrato[fk]
            ek  = fk

        if ek:
            processed_ek.add(ek)

        sem_fatura  = fat is None
        sem_extrato = ext is None
        nome_exibir = (fat or {}).get("nome_original") or (ext or {}).get("nome_original") or fk.title()

        val_esperado   = calc_esperado(fat, ext)
        val_descontado = (ext or {}).get("plano_descontado", 0.0)
        diferenca      = round(val_esperado - val_descontado, 2)

        if abs(diferenca) <= 0.05:
            status = "OK"
        elif diferenca > 0:
            status = "MAIOR"
        else:
            status = "MENOR"

        if status != "OK" or sem_fatura or sem_extrato:
            divergentes += 1

        total_esperado += val_esperado
        total_extrato  += val_descontado

        resultados.append({
            "nome":                   nome_exibir,
            "mensalidade_titular":    (fat or {}).get("mensalidade", 0.0),
            "mensalidade_dependentes":(fat or {}).get("mensalidade_dependentes", 0.0),
            "sos_tam":                (fat or {}).get("sos_tam", 0.0),
            "total_fatura":           (fat or {}).get("total", 0.0),
            "salario":                (ext or {}).get("salario", 0.0),
            "valor_esperado":         val_esperado,
            "dependentes":            (fat or {}).get("dependentes", 0),
            "valor_descontado":       val_descontado,
            "diferenca":              diferenca,
            "status":                 status,
            "sem_fatura":             sem_fatura,
            "sem_extrato":            sem_extrato,
        })

    # Funcionários só no extrato, não casados com fatura (evita duplicatas)
    for ek in extrato:
        if ek in processed_ek:
            continue  # já processado no loop principal
        if rev_map.get(ek) is None and ek not in fatura:
            ext = extrato[ek]
            val_descontado = ext.get("plano_descontado", 0.0)
            val_esperado   = calc_esperado(None, ext)
            diferenca      = round(val_esperado - val_descontado, 2)
            total_esperado += val_esperado
            total_extrato  += val_descontado
            divergentes    += 1
            resultados.append({
                "nome": ext.get("nome_original", ek.title()),
                "mensalidade_titular": 0.0, "mensalidade_dependentes": 0.0,
                "sos_tam": 0.0, "total_fatura": 0.0,
                "salario": ext.get("salario", 0.0),
                "valor_esperado": val_esperado,
                "dependentes": 0,
                "valor_descontado": val_descontado,
                "diferenca": diferenca,
                "status": "MENOR" if val_descontado > 0 else "OK",
                "sem_fatura": True, "sem_extrato": False,
            })

    resultados.sort(key=lambda x: (x["status"] == "OK", x["nome"]))

    return {
        "resultados":      resultados,
        "total_esperado":  round(total_esperado, 2),
        "total_extrato":   round(total_extrato, 2),
        "total_diferenca": round(total_esperado - total_extrato, 2),
        "divergentes":     divergentes,
        "total":           len(resultados),
        "regra":           regra,
    }


# ─────────────────────────────────────────────
# ROTAS
# ─────────────────────────────────────────────

@app.route("/")
def index():
    return render_template_string(HTML)

@app.route("/comparar-beneficio", methods=["POST"])
def comparar_beneficio():
    errors = []
    fatura_data = {}
    extrato_data = {}

    # Múltiplos arquivos de referência
    fatura_files = request.files.getlist("fatura")
    extrato_files = request.files.getlist("extrato")

    # Regra de desconto esperado
    regra_tipo  = request.form.get("regra_tipo", "fatura")
    regra_valor = float(request.form.get("regra_valor", 0) or 0)
    regra = {"tipo": regra_tipo, "valor": regra_valor}

    # Código do evento a analisar no extrato
    evento_codigo = request.form.get("evento_codigo", "8111").strip() or "8111"

    # Palavra-chave do documento de referência (default: MENSALIDADE)
    filtro_linha = request.form.get("filtro_linha", "MENSALIDADE").strip() or "MENSALIDADE"

    if not extrato_files or not any(f.filename for f in extrato_files):
        return jsonify({"error": "Envie ao menos o Extrato de Folha PDF."})

    for f in fatura_files:
        if not f.filename:
            continue
        try:
            fn = f.filename.lower()
            raw = f.read()
            if fn.endswith((".xls", ".xlsx")):
                fatura_data = _merge_fatura(fatura_data, parse_referencia_simples(raw, f.filename))
            else:
                fatura_data = _merge_fatura(fatura_data, parse_plano_fatura(raw, filtro_linha))
        except Exception as e:
            errors.append(f"{f.filename}: {e}")

    for f in extrato_files:
        if not f.filename:
            continue
        try:
            extrato_data = _merge_extrato(extrato_data, parse_extrato_plano(f.read(), evento_codigo))
        except Exception as e:
            errors.append(f"{f.filename}: {e}")

    if not extrato_data:
        return jsonify({"error": "Nenhum dado extraído do extrato. " + " | ".join(errors)})

    result = compare_plano_saude(fatura_data, extrato_data, regra)
    result["erros"] = errors
    # Debug: valores extraídos de cada arquivo
    result["debug_fatura"] = {
        k: {"nome": v.get("nome_original"), "valor": v.get("total", 0)}
        for k, v in fatura_data.items()
    }
    result["debug_extrato"] = {
        k: {"nome": v.get("nome_original"), "valor_descontado": v.get("plano_descontado", 0), "salario": v.get("salario", 0)}
        for k, v in extrato_data.items()
    }
    result["evento_codigo"] = evento_codigo
    return jsonify(result)

@app.route("/analisar", methods=["POST"])
def analisar():
    excel_data, pdf_data, word_data = {}, {}, {}
    errors = []

    for key, f in request.files.items():
        if not f.filename:
            continue
        raw = f.read()
        fn  = f.filename.lower()
        try:
            if fn.endswith((".xls", ".xlsx")):
                excel_data.update(parse_excel(raw, f.filename))
            elif fn.endswith(".pdf"):
                pdf_data.update(parse_pdf(raw))
            elif fn.endswith((".docx", ".doc")):
                parsed = parse_word(raw)
                word_data.setdefault("gratificacoes", {}).update(parsed.get("gratificacoes", {}))
                word_data.setdefault("comissao_e_dsr", {}).update(parsed.get("comissao_e_dsr", {}))
                word_data.setdefault("descontos", {}).update(parsed.get("descontos", {}))
                word_data.setdefault("obs", []).extend(parsed.get("obs", []))
                word_data.setdefault("decimo_terceiro", []).extend(parsed.get("decimo_terceiro", []))
        except Exception as e:
            errors.append(f"{f.filename}: {e}")

    if not excel_data and not pdf_data:
        return jsonify({"error": "Nenhum arquivo processado. " + " | ".join(errors)})

    report = compare(excel_data, pdf_data, word_data)
    report["erros"] = errors
    return jsonify(report)

# ─────────────────────────────────────────────
# HTML TEMPLATE
# ─────────────────────────────────────────────

HTML = r"""<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Conferência de Folha — Sigma Contabilidade</title>
<link href="https://fonts.googleapis.com/css2?family=Montserrat:wght@400;500;600;700;800&family=Inter:wght@400;500;600;700&display=swap" rel="stylesheet">
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:'Inter',sans-serif;background:#f1f3f6;color:#333333;min-height:100vh}

header{background:linear-gradient(135deg,#A72C31 0%,#8B2227 100%);color:#fff;padding:1.25rem 2rem;display:flex;align-items:center;justify-content:space-between;box-shadow:0 4px 20px rgba(0,0,0,.3)}
.logo{font-size:1.35rem;font-weight:700;letter-spacing:-.5px}.logo span{color:#A72C31}
.subtitle{font-size:.8rem;color:#8892b0;margin-top:2px}

.container{max-width:1050px;margin:0 auto;padding:2rem 1rem}

.card{background:#fff;border-radius:16px;padding:1.75rem;box-shadow:0 2px 16px rgba(0,0,0,.07);margin-bottom:1.25rem}

h2{font-size:1.05rem;font-weight:600;color:#A72C31;margin-bottom:.35rem}
.card > p{font-size:.83rem;color:#6b7280;margin-bottom:1.5rem}

.file-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(240px,1fr));gap:1rem;margin-bottom:1.5rem}

.zone{border:2px dashed #d1d5db;border-radius:12px;padding:1.5rem 1rem;text-align:center;cursor:pointer;transition:all .2s;position:relative;min-height:130px;display:flex;flex-direction:column;align-items:center;justify-content:center}
.zone:hover,.zone.over{border-color:#A72C31;background:#fff5f6}
.zone.done{border-color:#10b981;border-style:solid;background:#f0fdf4}
.zone input{position:absolute;inset:0;opacity:0;cursor:pointer;width:100%;height:100%}
.zone .icon{font-size:2rem;margin-bottom:.4rem}
.zone .lbl{font-size:.88rem;font-weight:600;color:#374151}
.zone .hint{font-size:.73rem;color:#9ca3af;margin-top:.2rem}
.zone .fname{font-size:.75rem;color:#059669;font-weight:500;margin-top:.4rem;word-break:break-all;max-width:200px}
.opt-badge{font-size:.65rem;color:#9ca3af;font-weight:400}

.btn{background:linear-gradient(135deg,#A72C31 0%,#8B2227 100%);color:#fff;border:none;padding:.85rem 2rem;border-radius:10px;font-size:.95rem;font-weight:600;cursor:pointer;transition:all .2s;width:100%;letter-spacing:.3px}
.btn:hover{transform:translateY(-1px);box-shadow:0 6px 20px rgba(233,69,96,.35)}
.btn:disabled{opacity:.6;cursor:not-allowed;transform:none}

.loading{display:none;text-align:center;padding:3rem}
.spinner{width:40px;height:40px;border:3px solid #f3f4f6;border-top-color:#A72C31;border-radius:50%;animation:spin .8s linear infinite;margin:0 auto 1rem}
@keyframes spin{to{transform:rotate(360deg)}}

.results{display:none}

.stats{display:grid;grid-template-columns:repeat(3,1fr);gap:1rem;margin-bottom:1.25rem}
.stat{background:#fff;border-radius:12px;padding:1.2rem 1rem;box-shadow:0 2px 10px rgba(0,0,0,.06);text-align:center}
.stat .n{font-size:2.4rem;font-weight:700;line-height:1}
.stat .l{font-size:.75rem;color:#6b7280;margin-top:.2rem;font-weight:500}
.stat.t .n{color:#333333}
.stat.g .n{color:#10b981}
.stat.d .n{color:#A72C31}

.sec-title{font-size:.95rem;font-weight:600;padding-bottom:.7rem;border-bottom:1px solid #f3f4f6;margin-bottom:1rem;display:flex;align-items:center;gap:.4rem}

.emp{border:1px solid #f0f0f0;border-radius:10px;margin-bottom:.6rem;overflow:hidden;transition:box-shadow .15s}
.emp:hover{box-shadow:0 2px 10px rgba(0,0,0,.07)}
.emp-hdr{display:flex;align-items:center;justify-content:space-between;padding:.8rem 1.1rem;cursor:pointer;transition:background .1s;gap:.5rem}
.emp-hdr:hover{background:#fafafa}
.emp-name{font-weight:600;font-size:.88rem}
.emp-sub{font-size:.75rem;color:#6b7280;margin-top:1px}
.badge{display:inline-flex;align-items:center;gap:.3rem;padding:.22rem .7rem;border-radius:20px;font-size:.72rem;font-weight:600;white-space:nowrap}
.badge.ok{background:#dcfce7;color:#166534}
.badge.div{background:#fee2e2;color:#991b1b}

.emp-body{display:none;padding:1rem 1.1rem;border-top:1px solid #f3f4f6;background:#fafafa}
.emp.open .emp-body{display:block}

.div-item{display:flex;gap:.7rem;padding:.55rem .8rem;border-radius:7px;margin-bottom:.45rem}
.div-item.alta{background:#fff1f2;border-left:3px solid #A72C31}
.div-item.media{background:#fffbeb;border-left:3px solid #f59e0b}
.div-item.manual{background:#eff6ff;border-left:3px solid #3b82f6}
.div-tipo{font-weight:600;font-size:.78rem;margin-bottom:.1rem}
.div-desc{font-size:.78rem;color:#4b5563}

.add-div-btn{margin-top:.7rem;background:none;border:1px dashed #d1d5db;border-radius:7px;padding:.4rem .8rem;font-size:.75rem;color:#6b7280;cursor:pointer;width:100%;text-align:left;transition:all .15s}
.add-div-btn:hover{border-color:#3b82f6;color:#3b82f6;background:#eff6ff}

.div-form{margin-top:.6rem;background:#f8faff;border:1px solid #dbeafe;border-radius:10px;padding:.9rem;display:none}
.div-form.open{display:block}
.div-form label{font-size:.75rem;font-weight:600;color:#374151;display:block;margin-bottom:.2rem;margin-top:.55rem}
.div-form label:first-child{margin-top:0}
.div-form input,.div-form textarea,.div-form select{width:100%;font-size:.8rem;padding:.42rem .6rem;border:1px solid #d1d5db;border-radius:6px;font-family:inherit;color:#111;background:#fff}
.div-form textarea{resize:vertical;min-height:52px}
.div-form .val-row{display:grid;grid-template-columns:1fr 1fr 1fr;gap:.5rem;align-items:end}
.div-form .diff-box{background:#e0edff;border-radius:6px;padding:.42rem .6rem;font-size:.8rem;font-weight:600;color:#1d4ed8;text-align:center}
.div-form .btn-row{display:flex;gap:.5rem;margin-top:.7rem}
.div-form .btn-salvar{background:#3b82f6;color:#fff;border:none;border-radius:7px;padding:.4rem 1rem;font-size:.78rem;font-weight:600;cursor:pointer}
.div-form .btn-salvar:hover{background:#2563eb}
.div-form .btn-cancel{background:none;border:1px solid #d1d5db;border-radius:7px;padding:.4rem .8rem;font-size:.78rem;color:#6b7280;cursor:pointer}
.div-form .btn-cancel:hover{border-color:#9ca3af;color:#374151}
.manual-tag{font-size:.65rem;background:#dbeafe;color:#1d4ed8;border-radius:4px;padding:.1rem .35rem;font-weight:600;margin-left:.3rem;vertical-align:middle}

.data-grid{display:grid;grid-template-columns:1fr 1fr;gap:1rem;margin-top:.75rem}
.dtbl h4{font-size:.78rem;font-weight:600;color:#374151;margin-bottom:.4rem}
.dtbl table{width:100%;border-collapse:collapse;font-size:.76rem}
.dtbl td{padding:.28rem .4rem;border-bottom:1px solid #f3f4f6}
.dtbl td:last-child{text-align:right;font-weight:500}
.dtbl tr:last-child td{font-weight:700;border-top:2px solid #e5e7eb;border-bottom:none}

details summary{font-size:.72rem;color:#6b7280;cursor:pointer;margin-top:.5rem;user-select:none}

.obs-box{background:#fffbeb;border-left:3px solid #f59e0b;border-radius:6px;padding:.5rem .8rem;font-size:.82rem;margin-bottom:.4rem}

.err-box{background:#fff1f2;border:1px solid #fecaca;border-radius:8px;padding:.9rem;margin-top:1rem}
.err-box h4{color:#b91c1c;font-size:.82rem;margin-bottom:.4rem}
.err-box p{font-size:.78rem;color:#7f1d1d}

.gratif-table{width:100%;border-collapse:collapse;font-size:.82rem;margin-top:.5rem}
.gratif-table th{text-align:left;padding:.3rem .5rem;color:#6b7280;font-weight:500;border-bottom:1px solid #f3f4f6}
.gratif-table td{padding:.3rem .5rem;border-bottom:1px solid #f9fafb}

@media(max-width:600px){
  .stats{grid-template-columns:1fr}
  .data-grid{grid-template-columns:1fr}
  .file-grid{grid-template-columns:1fr}
}

/* ── Tabs ── */
.tabs{display:flex;gap:.5rem;margin-bottom:1.25rem}
.tab{background:#fff;border:1px solid #e5e7eb;border-radius:10px;padding:.6rem 1.3rem;font-size:.85rem;font-weight:600;cursor:pointer;color:#6b7280;transition:all .2s}
.tab.active{background:#A72C31;color:#fff;border-color:#A72C31}
.section{display:none}
.section.active{display:block}

/* ── Tabela de Benefício ── */
.ben-table{width:100%;border-collapse:collapse;font-size:.82rem;table-layout:fixed}
.ben-table th{text-align:left;padding:.45rem .6rem;color:#6b7280;font-weight:500;border-bottom:2px solid #f0f0f0;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
.ben-table td{padding:.45rem .6rem;border-bottom:1px solid #f9fafb;vertical-align:middle;overflow:hidden;text-overflow:ellipsis}
.ben-table tr:hover td{background:#fafafa}
.ben-table tr.maior td{background:#fff1f2}
.ben-table tr.menor td{background:#fffbeb}
.ben-table tr.sem-doc td{background:#f0f9ff}
.ben-table .valor{text-align:right;font-weight:500;white-space:nowrap}
.ben-table .diff-pos{color:#A72C31;font-weight:700}
.ben-table .diff-neg{color:#f59e0b;font-weight:700}
.ben-table .diff-ok{color:#10b981;font-weight:700}

@page{margin:1cm 1.5cm;size:A4}
@media print{
  body > *:not(#print-overlay){display:none!important}
  #print-overlay{display:block!important;position:static!important;padding:0!important;overflow:visible!important}
}
.print-header{display:none;text-align:center;margin-bottom:1.2rem;font-size:.9rem;color:#333333}
.ben-badge{display:inline-flex;align-items:center;padding:.18rem .6rem;border-radius:20px;font-size:.7rem;font-weight:700;white-space:nowrap}
.ben-badge.ok{background:#dcfce7;color:#166534}
.ben-badge.maior{background:#fee2e2;color:#991b1b}
.ben-badge.menor{background:#fef9c3;color:#92400e}
.ben-badge.nd{background:#e0f2fe;color:#0369a1}
</style>
</head>
<body>

<header>
  <div style="display:flex;align-items:center;gap:1rem">
    <img src="https://gsigma.com.br/wp-content/uploads/2025/08/LOGO-SIGMA-ICONE.webp" alt="Sigma" style="height:40px;width:auto;filter:brightness(0) invert(1)">
    <div>
      <div class="logo" style="font-family:'Montserrat',sans-serif;font-size:1.3rem;font-weight:800;letter-spacing:-.5px;color:#fff">SIGMA <span style="color:#fff;font-weight:400;opacity:.85">Contabilidade</span></div>
      <div class="subtitle" style="font-size:.72rem;color:rgba(255,255,255,.65);letter-spacing:.3px">Além da Contabilidade</div>
    </div>
  </div>
  <div style="text-align:center">
    <div style="font-size:.95rem;font-weight:600;color:rgba(255,255,255,.9);letter-spacing:.2px">Conferência de Folha de Pagamento</div>
  </div>
  <div style="font-size:.72rem;color:rgba(255,255,255,.5);text-align:right;line-height:1.5">
    Arquivos processados<br>localmente · sem armazenamento
  </div>
</header>

<div class="container">

  <!-- TABS -->
  <div class="tabs no-print">
    <button class="tab active" onclick="switchTab('folha')">📊 Conferência de Folha</button>
    <button class="tab" onclick="switchTab('beneficio')">🏥 Conferência de Benefício</button>
  </div>

  <!-- SEÇÃO: CONFERÊNCIA DE FOLHA -->
  <div class="section active" id="section-folha">

    <!-- UPLOAD CARD -->
    <div class="card">
      <h2>Envie os arquivos da folha</h2>
      <p>Selecione ou arraste os arquivos abaixo. O sistema compara a planilha, os recibos e o documento de instruções e aponta todas as divergências.</p>

      <div class="file-grid">
        <div class="zone" id="zone-excel" ondragover="drag(event,'excel')" ondragleave="undrag(event,'excel')" ondrop="drop(event,'excel')">
          <input type="file" id="file-excel" accept=".xls,.xlsx" multiple onchange="sel('excel',this)">
          <div class="icon">📊</div>
          <div class="lbl">Planilha Excel</div>
          <div class="hint">.xls ou .xlsx — pode selecionar vários</div>
          <div class="fname" id="fname-excel"></div>
        </div>

        <div class="zone" id="zone-pdf" ondragover="drag(event,'pdf')" ondragleave="undrag(event,'pdf')" ondrop="drop(event,'pdf')">
          <input type="file" id="file-pdf" accept=".pdf" multiple onchange="sel('pdf',this)">
          <div class="icon">📄</div>
          <div class="lbl">Recibos PDF</div>
          <div class="hint">.pdf — pode selecionar vários</div>
          <div class="fname" id="fname-pdf"></div>
        </div>

        <div class="zone" id="zone-word" ondragover="drag(event,'word')" ondragleave="undrag(event,'word')" ondrop="drop(event,'word')">
          <input type="file" id="file-word" accept=".docx,.doc" multiple onchange="sel('word',this)">
          <div class="icon">📝</div>
          <div class="lbl">Documento Word <span class="opt-badge">(opcional)</span></div>
          <div class="hint">.docx — pode selecionar vários</div>
          <div class="fname" id="fname-word"></div>
        </div>
      </div>

      <button class="btn" id="btn" onclick="analyze()">🔍 Analisar Folha</button>
    </div>

    <!-- LOADING -->
    <div class="loading" id="loading">
      <div class="spinner"></div>
      <p style="color:#6b7280;font-size:.88rem">Processando e comparando dados...</p>
    </div>

    <!-- RESULTS -->
    <div class="results" id="results"></div>

  </div><!-- /section-folha -->

  <!-- SEÇÃO: CONFERÊNCIA DE BENEFÍCIO -->
  <div class="section" id="section-beneficio">

    <div class="card no-print">
      <h2>Conferência de Benefício</h2>
      <p>Compare o valor esperado de um benefício com o que foi descontado na folha de pagamento. Funciona com qualquer tipo de desconto: plano de saúde, vale alimentação, empréstimo, etc.</p>

      <div class="file-grid">
        <div class="zone" id="zone-fatura" ondragover="drag(event,'fatura')" ondragleave="undrag(event,'fatura')" ondrop="drop(event,'fatura')" style="position:relative">
          <input type="file" id="file-fatura" accept=".pdf" multiple onchange="sel('fatura',this)">
          <div class="icon">📁</div>
          <div class="lbl">Documento(s) de referência <span class="opt-badge">(opcional)</span></div>
          <div class="hint">.pdf — fatura, relatório, lista de valores — pode selecionar vários</div>
          <div class="fname" id="fname-fatura"></div>
        </div>

        <div style="display:flex;flex-direction:column;justify-content:center;gap:.5rem;padding:.5rem 0">
          <label style="font-size:.75rem;font-weight:600;color:#374151">Palavra-chave do documento PDF</label>
          <input type="text" id="filtro-linha" value="MENSALIDADE" placeholder="Ex: MENSALIDADE"
            style="font-size:.82rem;padding:.42rem .6rem;border:1px solid #d1d5db;border-radius:6px;width:100%;text-transform:uppercase">
          <div style="font-size:.7rem;color:#9ca3af">Palavra que identifica linhas de valor no PDF (padrão: MENSALIDADE)</div>
        </div>

        <div class="zone" id="zone-extrato" ondragover="drag(event,'extrato')" ondragleave="undrag(event,'extrato')" ondrop="drop(event,'extrato')">
          <input type="file" id="file-extrato" accept=".pdf" multiple onchange="sel('extrato',this)">
          <div class="icon">📋</div>
          <div class="lbl">Extrato de Folha PDF</div>
          <div class="hint">.pdf — pode selecionar vários</div>
          <div class="fname" id="fname-extrato"></div>
        </div>
      </div>

      <!-- Evento a analisar -->
      <div style="margin-bottom:1rem;background:#f0fdf4;border:1px solid #bbf7d0;border-radius:10px;padding:1rem">
        <div style="font-size:.82rem;font-weight:600;color:#166534;margin-bottom:.7rem">Evento a analisar no extrato</div>
        <div style="display:grid;grid-template-columns:1fr 1fr;gap:.8rem;align-items:end">
          <div>
            <label style="font-size:.75rem;font-weight:600;color:#374151;display:block;margin-bottom:.3rem">Evento</label>
            <select id="evento-select" onchange="toggleEventoCodigo()" style="width:100%;font-size:.82rem;padding:.42rem .6rem;border:1px solid #d1d5db;border-radius:6px;background:#fff">
              <option value="8111">8111 — Plano de Saúde</option>
              <option value="246">246 — Plano Odontológico</option>
              <option value="48">48 — Vale Transporte</option>
              <option value="203">203 — Alimentação</option>
              <option value="998">998 — INSS</option>
              <option value="999">999 — IRRF / Imposto de Renda</option>
              <option value="custom">Outro (digitar código)</option>
            </select>
          </div>
          <div id="evento-codigo-wrap" style="display:none">
            <label style="font-size:.75rem;font-weight:600;color:#374151;display:block;margin-bottom:.3rem">Código do evento</label>
            <input type="text" id="evento-codigo-input" placeholder="Ex: 1234" style="width:100%;font-size:.82rem;padding:.42rem .6rem;border:1px solid #d1d5db;border-radius:6px">
          </div>
          <div id="evento-info" style="font-size:.72rem;color:#166534;align-self:center">
            O sistema irá buscar e somar todos os descontos deste código no extrato para cada funcionário.
          </div>
        </div>
      </div>

      <!-- Regra de desconto esperado -->
      <div style="margin-bottom:1.2rem;background:#f8faff;border:1px solid #dbeafe;border-radius:10px;padding:1rem">
        <div style="font-size:.82rem;font-weight:600;color:#1e40af;margin-bottom:.7rem">Regra de desconto esperado</div>
        <div style="display:grid;grid-template-columns:1fr 1fr;gap:.8rem;align-items:end">
          <div>
            <label style="font-size:.75rem;font-weight:600;color:#374151;display:block;margin-bottom:.3rem">Tipo</label>
            <select id="regra-tipo" onchange="toggleRegraValor()" style="width:100%;font-size:.82rem;padding:.42rem .6rem;border:1px solid #d1d5db;border-radius:6px;background:#fff">
              <option value="fatura">Usar valor do documento de referência</option>
              <option value="pct_fatura">% do valor do documento de referência</option>
              <option value="pct_salario">% do salário do funcionário</option>
              <option value="fixo">Valor fixo por funcionário</option>
            </select>
          </div>
          <div id="regra-valor-wrap">
            <label style="font-size:.75rem;font-weight:600;color:#374151;display:block;margin-bottom:.3rem" id="regra-valor-label">Percentual ou valor</label>
            <input type="number" id="regra-valor" step="0.01" min="0" placeholder="Ex: 15 para 15%" style="width:100%;font-size:.82rem;padding:.42rem .6rem;border:1px solid #d1d5db;border-radius:6px">
          </div>
        </div>
        <div id="regra-hint" style="font-size:.72rem;color:#6b7280;margin-top:.5rem"></div>
      </div>

      <button class="btn" id="btn-beneficio" onclick="analyzeBeneficio()">🔍 Analisar Benefício</button>
    </div>

    <!-- LOADING BENEFÍCIO -->
    <div class="loading" id="loading-beneficio">
      <div class="spinner"></div>
      <p style="color:#6b7280;font-size:.88rem">Processando faturas e extraindo descontos...</p>
    </div>

    <!-- RESULTS BENEFÍCIO -->
    <div id="results-beneficio"></div>

  </div><!-- /section-beneficio -->

</div>

<script>
const FILES = {excel:[], pdf:[], word:[], fatura:[], extrato:[]};

function switchTab(tab) {
  document.querySelectorAll('.tab').forEach(t => t.classList.remove('active'));
  document.querySelectorAll('.section').forEach(s => s.classList.remove('active'));
  document.querySelector(`.tab[onclick="switchTab('${tab}')"]`).classList.add('active');
  document.getElementById(`section-${tab}`).classList.add('active');
}

function sel(type, inp) {
  FILES[type] = Array.from(inp.files);
  const names = FILES[type].map(f=>f.name).join(' • ');
  document.getElementById(`fname-${type}`).textContent = names;
  document.getElementById(`zone-${type}`).classList.toggle('done', FILES[type].length>0);
}

function toggleEventoCodigo(){
  const sel = document.getElementById('evento-select');
  const wrap = document.getElementById('evento-codigo-wrap');
  const info = document.getElementById('evento-info');
  const isCustom = sel.value === 'custom';
  wrap.style.display = isCustom ? 'block' : 'none';
  info.style.display = isCustom ? 'none' : 'block';
}

function getEventoCodigo(){
  const sel = document.getElementById('evento-select');
  if(sel.value === 'custom'){
    return document.getElementById('evento-codigo-input').value.trim() || '8111';
  }
  return sel.value;
}

function toggleRegraValor(){
  const tipo = document.getElementById('regra-tipo').value;
  const wrap = document.getElementById('regra-valor-wrap');
  const label = document.getElementById('regra-valor-label');
  const hint  = document.getElementById('regra-hint');
  const inp   = document.getElementById('regra-valor');
  if(tipo === 'fatura'){
    wrap.style.opacity='0.4'; inp.disabled=true;
    hint.textContent='O valor esperado será lido diretamente do documento de referência.';
  } else {
    wrap.style.opacity='1'; inp.disabled=false;
    if(tipo==='pct_fatura'){label.textContent='Percentual (%)';inp.placeholder='Ex: 15';hint.textContent='Calcula X% do valor do documento de referência por funcionário.';}
    else if(tipo==='pct_salario'){label.textContent='Percentual (%)';inp.placeholder='Ex: 15';hint.textContent='Calcula X% do salário de cada funcionário (requer campo Salário no extrato).';}
    else if(tipo==='fixo'){label.textContent='Valor fixo (R$)';inp.placeholder='Ex: 150.00';hint.textContent='Aplica o mesmo valor esperado para todos os funcionários.';}
  }
}

function drag(e,t){e.preventDefault();document.getElementById(`zone-${t}`).classList.add('over')}
function undrag(e,t){document.getElementById(`zone-${t}`).classList.remove('over')}
function drop(e,t){
  e.preventDefault();
  document.getElementById(`zone-${t}`).classList.remove('over');
  const inp=document.getElementById(`file-${t}`);
  const dt=new DataTransfer();
  Array.from(e.dataTransfer.files).forEach(f=>dt.items.add(f));
  inp.files=dt.files;
  sel(t,inp);
}

async function analyze(){
  if(!FILES.excel.length && !FILES.pdf.length){alert('Envie pelo menos a planilha Excel ou os recibos PDF.');return}
  const btn=document.getElementById('btn');
  btn.disabled=true;
  document.getElementById('loading').style.display='block';
  document.getElementById('results').style.display='none';

  const fd=new FormData();
  FILES.excel.forEach((f,i)=>fd.append(`excel_${i}`,f));
  FILES.pdf.forEach((f,i)=>fd.append(`pdf_${i}`,f));
  FILES.word.forEach((f,i)=>fd.append(`word_${i}`,f));

  try{
    const res=await fetch('/analisar',{method:'POST',body:fd});
    const data=await res.json();
    render(data);
  } catch(e){
    show(`<div class="err-box"><h4>Erro de comunicação</h4><p>${e.message}</p></div>`);
  } finally{
    btn.disabled=false;
    document.getElementById('loading').style.display='none';
  }
}

function show(html){
  const el=document.getElementById('results');
  el.style.display='block';
  el.innerHTML=html;
  el.scrollIntoView({behavior:'smooth'});
}

function brl(v){
  if(!v&&v!==0)return'-';
  return'R$ '+Number(v).toLocaleString('pt-BR',{minimumFractionDigits:2,maximumFractionDigits:2});
}

function render(data){
  if(data.error){show(`<div class="err-box"><h4>Erro</h4><p>${data.error}</p></div>`);return}

  const r=data.resumo;
  let html=`
    <div class="stats">
      <div class="stat t"><div class="n">${r.total}</div><div class="l">Funcionários analisados</div></div>
      <div class="stat g"><div class="n">${r.ok}</div><div class="l">Sem divergências</div></div>
      <div class="stat d"><div class="n">${r.divergencias}</div><div class="l">Com divergências</div></div>
    </div>`;

  // Gratificações do Word
  const wg=data.word_gratificacoes||{};
  if(Object.keys(wg).length){
    html+=`<div class="card"><div class="sec-title">💰 Gratificações informadas no Word</div>
      <table class="gratif-table">
        <tr><th>Funcionário</th><th>Valor</th><th>No recibo?</th></tr>
        ${Object.entries(wg).map(([n,v])=>{
          const rec=data.funcionarios.find(f=>f.nome===n);
          const ok=rec?.dados_recibo?.has_gratif;
          return `<tr><td>${(rec?.nome_exibir||n)}</td><td>${v}</td>
            <td>${ok?'<span style="color:#059669;font-weight:600">✓ Sim</span>':'<span style="color:#A72C31;font-weight:600">✗ Não</span>'}</td></tr>`;
        }).join('')}
      </table></div>`;
  }

  // Observações
  if(data.observacoes?.length){
    html+=`<div class="card"><div class="sec-title">⚠️ Observações do documento Word</div>
      ${data.observacoes.map(o=>`<div class="obs-box">${o}</div>`).join('')}
    </div>`;
  }

  // Funcionários
  html+=`<div class="card"><div class="sec-title">👥 Resultado por funcionário</div>`;

  const sorted=[...data.funcionarios].sort((a,b)=>{
    if(a.status!==b.status) return a.status==='DIVERGENTE'?-1:1;
    return a.nome.localeCompare(b.nome);
  });

  for(const emp of sorted){
    const divg=emp.status==='DIVERGENTE';
    const badge=divg
      ?`<span class="badge div">⚠ ${emp.divs.length} divergência${emp.divs.length>1?'s':''}</span>`
      :`<span class="badge ok">✓ OK</span>`;

    const recTipo={mensal:'Mensal','13_adiantamento':'13º Adiant.',ferias:'Férias'}[emp.dados_recibo?.tipo]||'';

    html+=`<div class="emp${divg?' open':''}" onclick="tog(this)">
      <div class="emp-hdr">
        <div style="flex:1;min-width:0">
          <div class="emp-name">${emp.nome_exibir||emp.nome}</div>
          <div class="emp-sub">
            ${emp.dados_excel?`Planilha: ${brl(emp.dados_excel.liquido)}`:'Sem planilha'}
            ${emp.dados_recibo?` · Recibo: ${brl(emp.dados_recibo.liquido)} ${recTipo?'('+recTipo+')':''}`:'· Sem recibo'}
          </div>
        </div>
        ${badge}
      </div>`;

    if(divg||emp.dados_excel||emp.dados_recibo){
      html+=`<div class="emp-body">`;

      // Divergências
      for(const d of emp.divs){
        html+=`<div class="div-item ${d.g}"><div><div class="div-tipo">${d.tipo}</div><div class="div-desc">${d.desc}</div></div></div>`;
      }

      // Tabelas de dados (collapsível)
      if(emp.dados_excel || emp.dados_recibo){
        html+=`<details style="margin-top:.6rem">
          <summary style="font-size:.75rem;color:#6b7280;cursor:pointer;user-select:none;padding:.25rem 0;list-style:none">
            <span style="text-decoration:underline;text-decoration-style:dotted">Ver dados completos (planilha / recibo)</span>
          </summary>
          <div class="data-grid" style="margin-top:.6rem">`;
        if(emp.dados_excel){
          const e=emp.dados_excel;
          html+=`<div class="dtbl"><h4>📊 Planilha Excel</h4><table>
            <tr><td>Salário</td><td>${brl(e.salario)}</td></tr>
            ${e.gratificacao?`<tr><td>Gratificação</td><td>${brl(e.gratificacao)}</td></tr>`:''}
            ${e.ferias_13?`<tr><td>Férias / 13º</td><td>${brl(e.ferias_13)}</td></tr>`:''}
            ${e.inss?`<tr><td>INSS</td><td>− ${brl(e.inss)}</td></tr>`:''}
            ${e.vale?`<tr><td>Vale Transp.</td><td>− ${brl(e.vale)}</td></tr>`:''}
            ${e.plano?`<tr><td>Plano/Unimed</td><td>− ${brl(e.plano)}</td></tr>`:''}
            ${e.emprestimo?`<tr><td>Empréstimo</td><td>− ${brl(e.emprestimo)}</td></tr>`:''}
            <tr><td>Líquido</td><td>${brl(e.liquido)}</td></tr>
          </table></div>`;
        }
        if(emp.dados_recibo){
          const r=emp.dados_recibo;
          html+=`<div class="dtbl"><h4>📄 Recibo PDF</h4><table>
            <tr><td>Total Vencimentos</td><td>${brl(r.total_vencimentos)}</td></tr>
            <tr><td>Total Descontos</td><td>− ${brl(r.total_descontos)}</td></tr>
            <tr><td>Líquido</td><td>${brl(r.liquido)}</td></tr>
          </table>
          ${r.verbas?.length?`<details><summary style="font-size:.72rem;cursor:pointer;color:#6b7280">Ver ${r.verbas.length} verbas</summary>
            <table style="margin-top:.4rem;font-size:.73rem">
              ${r.verbas.map(v=>`<tr><td>${v.codigo} — ${v.descricao}</td><td style="text-align:right;padding-left:.5rem">${brl(v.valor)}</td></tr>`).join('')}
            </table></details>`:''}
          </div>`;
        }
        html+=`</div></details>`;
      }
      // Formulário de divergência manual
      html+=`
        <button class="add-div-btn" onclick="toggleForm(event,this)">+ Apontar divergência manual</button>
        <div class="div-form" id="form-${emp.nome.replace(/\s+/g,'_')}">
          <label>Tipo de divergência</label>
          <input list="tipos-list" placeholder="Ex: Plano de saúde, INSS, Vale transporte..." class="f-tipo">
          <datalist id="tipos-list">
            <option value="Plano de saúde descontado incorretamente">
            <option value="INSS calculado errado">
            <option value="Vale transporte incorreto">
            <option value="IRRF divergente">
            <option value="Gratificação ausente">
            <option value="Horas extras não pagas">
            <option value="Desconto indevido">
            <option value="Adiantamento não descontado">
            <option value="Salário divergente">
          </datalist>
          <label>Descrição (explique o problema)</label>
          <textarea class="f-desc" placeholder="Descreva a divergência encontrada..."></textarea>
          <label>Valores</label>
          <div class="val-row">
            <div>
              <div style="font-size:.7rem;color:#6b7280;margin-bottom:.2rem">Esperado</div>
              <input type="number" step="0.01" placeholder="0,00" class="f-esp" oninput="calcDiff(this)">
            </div>
            <div>
              <div style="font-size:.7rem;color:#6b7280;margin-bottom:.2rem">Encontrado</div>
              <input type="number" step="0.01" placeholder="0,00" class="f-enc" oninput="calcDiff(this)">
            </div>
            <div>
              <div style="font-size:.7rem;color:#6b7280;margin-bottom:.2rem">Diferença</div>
              <div class="diff-box f-diff">—</div>
            </div>
          </div>
          <div class="btn-row">
            <button class="btn-salvar" onclick="saveDiv(this,'${emp.nome_exibir||emp.nome}')">Salvar divergência</button>
            <button class="btn-cancel" onclick="toggleForm(event,this.closest('.div-form').previousElementSibling)">Cancelar</button>
          </div>
        </div>`;

      html+=`</div>`;
    }
    html+=`</div>`;
  }

  html+=`</div>`;

  // Erros de processamento
  if(data.erros?.length){
    html+=`<div class="err-box"><h4>Avisos de processamento</h4>${data.erros.map(e=>`<p>${e}</p>`).join('')}</div>`;
  }

  // Botão imprimir
  html+=`<div style="text-align:center;margin-top:1rem">
    <button onclick="window.print()" style="background:none;border:1px solid #d1d5db;padding:.5rem 1.5rem;border-radius:8px;font-size:.82rem;cursor:pointer;color:#6b7280">🖨 Imprimir / Salvar PDF</button>
  </div>`;

  show(html);
}

async function analyzeBeneficio(){
  if(!FILES.extrato.length){alert('Envie o Extrato de Folha PDF.');return}
  const btn=document.getElementById('btn-beneficio');
  btn.disabled=true;
  document.getElementById('loading-beneficio').style.display='block';
  document.getElementById('results-beneficio').innerHTML='';

  const fd=new FormData();
  FILES.fatura.forEach(f=>fd.append('fatura',f));
  FILES.extrato.forEach(f=>fd.append('extrato',f));
  fd.append('regra_tipo', document.getElementById('regra-tipo').value);
  fd.append('regra_valor', document.getElementById('regra-valor').value||'0');
  fd.append('evento_codigo', getEventoCodigo());
  fd.append('filtro_linha', (document.getElementById('filtro-linha').value||'MENSALIDADE').toUpperCase());

  try{
    const res=await fetch('/comparar-beneficio',{method:'POST',body:fd});
    const data=await res.json();
    renderBeneficio(data);
  } catch(e){
    document.getElementById('results-beneficio').innerHTML=
      `<div class="err-box"><h4>Erro de comunicação</h4><p>${e.message}</p></div>`;
  } finally{
    btn.disabled=false;
    document.getElementById('loading-beneficio').style.display='none';
  }
}

function renderBeneficio(data){
  const el=document.getElementById('results-beneficio');
  if(data.error){
    el.innerHTML=`<div class="err-box"><h4>Erro</h4><p>${data.error}</p></div>`;
    return;
  }

  const difClass = v => v > 0.05 ? 'diff-pos' : (v < -0.05 ? 'diff-neg' : 'diff-ok');
  const difSign  = v => v > 0.05 ? '+' : '';
  const fmtDif   = v => {
    const cls=difClass(v);
    const sign=difSign(v);
    return `<span class="${cls}">${sign}${brl(v)}</span>`;
  };

  const regra = data.regra || {};
  const regraDesc = {
    fatura: 'Valor do documento de referência',
    pct_fatura: `${regra.valor}% do valor do documento`,
    pct_salario: `${regra.valor}% do salário`,
    fixo: `Valor fixo R$ ${brl(regra.valor)}`
  }[regra.tipo] || '';

  // Cards de resumo
  const difAbs = Math.abs(data.total_diferenca);
  const difLabel = data.total_diferenca > 0.05 ? 'A Descontar' : (data.total_diferenca < -0.05 ? 'A Devolver' : 'Tudo OK');
  const difColor = data.total_diferenca > 0.05 ? '#A72C31' : (data.total_diferenca < -0.05 ? '#f59e0b' : '#10b981');
  const difBg    = data.total_diferenca > 0.05 ? '#fff7ed' : (data.total_diferenca < -0.05 ? '#fffbeb' : '#f0fdf4');
  const okCount  = data.total - data.divergentes;

  let html=`<div style="display:grid;grid-template-columns:repeat(3,1fr);gap:1rem;margin-bottom:1.25rem">

    <div style="background:#fff;border-radius:12px;padding:1.2rem 1rem;box-shadow:0 2px 10px rgba(0,0,0,.06);text-align:center;border-top:3px solid #A72C31">
      <div style="font-size:2.4rem;font-weight:800;color:#A72C31;line-height:1;font-family:'Montserrat',sans-serif">${data.total}</div>
      <div style="font-size:.72rem;font-weight:600;color:#666;margin-top:.35rem;text-transform:uppercase;letter-spacing:.5px">Funcionários</div>
    </div>

    <div style="background:${difBg};border-radius:12px;padding:1.2rem 1rem;box-shadow:0 2px 10px rgba(0,0,0,.06);text-align:center;border-top:3px solid ${difColor}">
      <div style="font-size:.65rem;font-weight:700;color:${difColor};text-transform:uppercase;letter-spacing:.5px;margin-bottom:.25rem">${difLabel}</div>
      <div style="font-size:1.5rem;font-weight:800;color:${difColor};line-height:1">${difAbs > 0.05 ? brl(difAbs) : '—'}</div>
      <div style="font-size:.65rem;color:#9ca3af;margin-top:.25rem">diferença total</div>
    </div>

    <div style="background:#fff;border-radius:12px;padding:1.2rem 1rem;box-shadow:0 2px 10px rgba(0,0,0,.06);text-align:center;border-top:3px solid ${data.divergentes>0?'#A72C31':'#10b981'}">
      <div style="display:flex;gap:1.5rem;align-items:center;justify-content:center">
        <div><div style="font-size:2rem;font-weight:800;color:#A72C31;font-family:'Montserrat',sans-serif">${data.divergentes}</div><div style="font-size:.65rem;color:#666;font-weight:600;margin-top:.15rem">divergentes</div></div>
        <div style="color:#e5e7eb;font-size:1.2rem">|</div>
        <div><div style="font-size:2rem;font-weight:800;color:#10b981;font-family:'Montserrat',sans-serif">${okCount}</div><div style="font-size:.65rem;color:#666;font-weight:600;margin-top:.15rem">ok</div></div>
      </div>
    </div>

  </div>`;

  if(regraDesc) html+=`<div style="background:#fff7ed;border-radius:8px;padding:.5rem 1rem;font-size:.78rem;color:#92400e;margin-bottom:.8rem;border-left:3px solid #A72C31">Regra aplicada: <strong>${regraDesc}</strong> &nbsp;·&nbsp; Evento analisado: <strong>${data.evento_codigo||'8111'}</strong></div>`;

  // Seção de dados extraídos (debug / verificação)
  const fat = data.debug_fatura || {};
  const ext = data.debug_extrato || {};
  const fatKeys = Object.keys(fat);
  const extKeys = Object.keys(ext);
  if(fatKeys.length || extKeys.length){
    const allKeys = [...new Set([...fatKeys, ...extKeys])];
    const semMatch = allKeys.filter(k=>!(fat[k]&&ext[k])).length;
    html+=`<div style="background:#f0fdf4;border:1px solid #bbf7d0;border-radius:10px;padding:.75rem 1rem;margin-bottom:1rem;display:flex;align-items:center;justify-content:space-between;flex-wrap:wrap;gap:.5rem" class="no-print">
      <div style="font-size:.8rem;color:#166534">
        <strong>Leitura dos arquivos:</strong>
        &nbsp;Fatura: <strong>${fatKeys.length}</strong> registros
        &nbsp;·&nbsp; Extrato: <strong>${extKeys.length}</strong> registros
        ${semMatch>0 ? `&nbsp;·&nbsp; <span style="color:#dc2626;font-weight:700">${semMatch} sem correspondência</span>` : '&nbsp;·&nbsp; <span style="color:#166534">todos casados ✓</span>'}
      </div>
      <details style="font-size:.78rem">
        <summary style="cursor:pointer;color:#6b7280;user-select:none">Ver detalhes de extração</summary>
        <div style="overflow-x:auto;margin-top:.5rem;background:#fff;border-radius:8px;padding:.5rem">
        <table style="width:100%;border-collapse:collapse;font-size:.75rem">
          <thead><tr style="background:#f3f4f6">
            <th style="text-align:left;padding:.35rem .5rem;border-bottom:2px solid #e5e7eb">Funcionário</th>
            <th style="text-align:right;padding:.35rem .5rem;border-bottom:2px solid #e5e7eb">Valor na fatura</th>
            <th style="text-align:right;padding:.35rem .5rem;border-bottom:2px solid #e5e7eb">Descontado (evento ${data.evento_codigo||'8111'})</th>
            <th style="text-align:right;padding:.35rem .5rem;border-bottom:2px solid #e5e7eb">Salário</th>
            <th style="text-align:center;padding:.35rem .5rem;border-bottom:2px solid #e5e7eb">Encontrado em</th>
          </tr></thead>
          <tbody>
          ${allKeys.map(k=>{
            const f = fat[k];
            const e = ext[k];
            const nome = (f||e)?.nome || k;
            const vf = f ? brl(f.valor) : '—';
            const ve = e ? brl(e.valor_descontado) : '—';
            const sal = e ? brl(e.salario) : '—';
            const onde = (f&&e)?'✅ Ambos':(f?'⚠️ Só fatura':'⚠️ Só extrato');
            const cls = (f&&e)?'':'background:#fffbeb';
            return `<tr style="${cls}">
              <td style="padding:.3rem .5rem;border-bottom:1px solid #f3f4f6">${nome}</td>
              <td style="text-align:right;padding:.3rem .5rem;border-bottom:1px solid #f3f4f6;font-weight:500">${vf}</td>
              <td style="text-align:right;padding:.3rem .5rem;border-bottom:1px solid #f3f4f6;font-weight:500">${ve}</td>
              <td style="text-align:right;padding:.3rem .5rem;border-bottom:1px solid #f3f4f6;color:#6b7280">${sal}</td>
              <td style="text-align:center;padding:.3rem .5rem;border-bottom:1px solid #f3f4f6">${onde}</td>
            </tr>`;
          }).join('')}
          </tbody>
        </table>
        </div>
      </details>
    </div>`
  }

  // Resumo de pendências (só quem tem diferença)
  const aDescontar = data.resultados.filter(r=>r.status==='MAIOR' && Math.abs(r.diferenca)>0.05);
  const aDevolver  = data.resultados.filter(r=>r.status==='MENOR' && Math.abs(r.diferenca)>0.05);

  if(aDescontar.length || aDevolver.length){
    html+=`<div class="card" style="border-left:4px solid #A72C31">
      <div class="sec-title" style="margin-bottom:.8rem">⚠️ Pendências a resolver</div>
      <div style="display:grid;grid-template-columns:1fr 1fr;gap:1rem">`;

    if(aDescontar.length){
      const totalDesc = aDescontar.reduce((s,r)=>s+Math.abs(r.diferenca),0);
      html+=`<div>
        <div style="font-size:.78rem;font-weight:700;color:#A72C31;margin-bottom:.5rem">
          Faltou descontar — Total: <strong>${brl(totalDesc)}</strong>
        </div>
        <table style="width:100%;font-size:.78rem;border-collapse:collapse">
          ${aDescontar.map(r=>`
          <tr style="border-bottom:1px solid #f9fafb">
            <td style="padding:.3rem .2rem">${r.nome}</td>
            <td style="text-align:right;padding:.3rem .2rem;font-weight:700;color:#A72C31">
              Descontar ${brl(Math.abs(r.diferenca))}
            </td>
          </tr>`).join('')}
        </table>
      </div>`;
    }

    if(aDevolver.length){
      const totalDev = aDevolver.reduce((s,r)=>s+Math.abs(r.diferenca),0);
      html+=`<div>
        <div style="font-size:.78rem;font-weight:700;color:#f59e0b;margin-bottom:.5rem">
          Descontou a mais — Total: <strong>${brl(totalDev)}</strong>
        </div>
        <table style="width:100%;font-size:.78rem;border-collapse:collapse">
          ${aDevolver.map(r=>`
          <tr style="border-bottom:1px solid #f9fafb">
            <td style="padding:.3rem .2rem">${r.nome}</td>
            <td style="text-align:right;padding:.3rem .2rem;font-weight:700;color:#f59e0b">
              Devolver ${brl(Math.abs(r.diferenca))}
            </td>
          </tr>`).join('')}
        </table>
      </div>`;
    }

    html+=`</div></div>`;
  }

  // ── TABELA DE RESULTADOS ──────────────────────────────────────────

  const labelEsperado = regraDesc || 'Esperado';

  html+=`<div class="card">
    <div class="sec-title" style="margin-bottom:.75rem">Resultado por funcionário
      <span style="font-size:.72rem;font-weight:400;color:#9ca3af;margin-left:.5rem">· soma de todos os meses processados</span>
    </div>
    <div style="font-size:.72rem;color:#6b7280;display:flex;gap:1.2rem;flex-wrap:wrap;margin-bottom:1rem">
      <span><span class="ben-badge maior" style="font-size:.68rem">A DESCONTAR</span> faltou descontar</span>
      <span><span class="ben-badge menor" style="font-size:.68rem">A DEVOLVER</span> descontou a mais</span>
      <span><span class="ben-badge nd" style="font-size:.68rem">S/ DOC</span> sem correspondência</span>
    </div>
    <div style="overflow-x:auto">
    <table class="ben-table">
      <colgroup>
        <col style="width:32%"><col style="width:17%"><col style="width:17%"><col style="width:14%"><col style="width:20%">
      </colgroup>
      <thead><tr>
        <th>Funcionário</th>
        <th style="text-align:right">${labelEsperado}</th>
        <th style="text-align:right">Descontado</th>
        <th style="text-align:right">Diferença</th>
        <th style="text-align:center">Status</th>
      </tr></thead>
      <tbody>`;

  for(const r of data.resultados){
    const isOk = r.status==='OK' && !r.sem_extrato;
    const rowCls = r.sem_extrato?'sem-doc':(r.status==='MAIOR'?'maior':(r.status==='MENOR'?'menor':''));
    let badge = r.sem_extrato ? `<span class="ben-badge nd">S/ DOC</span>`
      : isOk ? `<span class="ben-badge ok">OK</span>`
      : r.status==='MAIOR' ? `<span class="ben-badge maior">A DESCONTAR</span>`
      : `<span class="ben-badge menor">A DEVOLVER</span>`;
    const detalhe = r.total_fatura ? `title="Titular: ${brl(r.mensalidade_titular)} | Dep.: ${brl(r.mensalidade_dependentes)} | SOS/TAM: ${brl(r.sos_tam)}"` : '';
    html+=`<tr class="${rowCls}">
      <td style="font-weight:500">${r.nome}${r.sem_extrato?'<br><span style="font-size:.68rem;color:#dc2626">sem extrato</span>':''}</td>
      <td class="valor" ${detalhe}>${r.valor_esperado?brl(r.valor_esperado):'—'}</td>
      <td class="valor">${r.valor_descontado?brl(r.valor_descontado):'—'}</td>
      <td class="valor">${fmtDif(r.diferenca)}</td>
      <td style="text-align:center">${badge}</td>
    </tr>`;
  }

  html+=`</tbody>
    <tfoot><tr style="background:#f1f3f6;font-weight:700;font-size:.85rem">
      <td style="padding:.55rem .7rem;border-top:2px solid #e5e7eb">TOTAL GERAL</td>
      <td style="text-align:right;padding:.55rem .7rem;border-top:2px solid #e5e7eb">${brl(data.total_esperado)}</td>
      <td style="text-align:right;padding:.55rem .7rem;border-top:2px solid #e5e7eb">${brl(data.total_extrato)}</td>
      <td style="text-align:right;padding:.55rem .7rem;border-top:2px solid #e5e7eb;color:${data.total_diferenca>0.05?'#A72C31':data.total_diferenca<-0.05?'#f59e0b':'#10b981'};font-weight:800">
        ${Math.abs(data.total_diferenca)>0.05?(data.total_diferenca>0?'+':'')+brl(data.total_diferenca):'—'}
      </td>
      <td style="padding:.55rem .7rem;border-top:2px solid #e5e7eb"></td>
    </tr></tfoot>
  </table>
  </div>
  </div>`;

  if(data.erros?.length){
    html+=`<div class="err-box"><h4>Avisos de processamento</h4>${data.erros.map(e=>`<p>${e}</p>`).join('')}</div>`;
  }
  html+=`<div class="no-print" style="text-align:center;margin-top:1.25rem">
    <button onclick="printRelatorio('${encodeURIComponent(regraDesc)}')" style="background:linear-gradient(135deg,#A72C31 0%,#8B2227 100%);color:#fff;border:none;padding:.8rem 2.5rem;border-radius:10px;font-size:.92rem;font-weight:600;cursor:pointer;letter-spacing:.3px;box-shadow:0 4px 14px rgba(167,44,49,.35)">Imprimir Relatório</button>
  </div>`;

  el.innerHTML=html;
  el.scrollIntoView({behavior:'smooth'});
}

function printRelatorio(regraEnc){
  const regra = decodeURIComponent(regraEnc||'');
  const hoje = new Date().toLocaleDateString('pt-BR',{day:'2-digit',month:'long',year:'numeric'});

  // Coleta apenas linhas com divergência (maior, menor, sem-doc)
  const rows = document.querySelectorAll('#results-beneficio .ben-table tbody tr.maior, #results-beneficio .ben-table tbody tr.menor, #results-beneficio .ben-table tbody tr.sem-doc');
  let tbodyHtml = '';
  rows.forEach(r => { tbodyHtml += r.outerHTML; });

  // Totais dos cards
  const statsEl = document.querySelector('#results-beneficio > div[style*="grid-template-columns"]');
  const statsHtml = statsEl ? statsEl.outerHTML : '';

  // Monta o HTML do relatório
  const reportHtml = `
    <div style="background:linear-gradient(135deg,#A72C31 0%,#8B2227 100%);padding:1.5rem 2rem;display:flex;align-items:center;justify-content:space-between;margin-bottom:1.4rem">
      <div style="display:flex;align-items:center;gap:1rem">
        <img src="https://gsigma.com.br/wp-content/uploads/2025/08/LOGO-SIGMA-ICONE.webp" style="height:44px;filter:brightness(0) invert(1)">
        <div>
          <div style="font-family:'Montserrat',sans-serif;font-size:1.3rem;font-weight:800;color:#fff;letter-spacing:-.4px;line-height:1.1">SIGMA <span style="font-weight:400;opacity:.8">Contabilidade</span></div>
          <div style="font-size:.62rem;color:rgba(255,255,255,.6);letter-spacing:.5px;margin-top:.1rem">ALÉM DA CONTABILIDADE</div>
        </div>
      </div>
      <div style="text-align:right">
        <div style="font-size:.9rem;font-weight:700;color:#fff">Conferência de Benefício</div>
        <div style="font-size:.7rem;color:rgba(255,255,255,.7);margin-top:.2rem">${hoje}</div>
        ${regra?`<div style="font-size:.65rem;color:rgba(255,255,255,.55);margin-top:.1rem">${regra}</div>`:''}
      </div>
    </div>
    ${statsHtml}
    <div style="font-size:.85rem;font-weight:700;color:#A72C31;margin:1.2rem 0 .5rem;padding-bottom:.4rem;border-bottom:2px solid #A72C31">
      Divergências encontradas
    </div>
    ${tbodyHtml?`<table style="width:100%;border-collapse:collapse;font-size:.8rem">
      <thead><tr style="background:#f9fafb">
        <th style="text-align:left;padding:.45rem .5rem;color:#555;border-bottom:2px solid #e5e7eb;font-weight:600">Funcionário</th>
        <th style="text-align:right;padding:.45rem .5rem;color:#555;border-bottom:2px solid #e5e7eb;font-weight:600">Esperado</th>
        <th style="text-align:right;padding:.45rem .5rem;color:#555;border-bottom:2px solid #e5e7eb;font-weight:600">Descontado</th>
        <th style="text-align:right;padding:.45rem .5rem;color:#555;border-bottom:2px solid #e5e7eb;font-weight:600">Diferença</th>
        <th style="text-align:center;padding:.45rem .5rem;color:#555;border-bottom:2px solid #e5e7eb;font-weight:600">Status</th>
      </tr></thead>
      <tbody>${tbodyHtml}</tbody>
    </table>`:'<p style="color:#666;font-size:.85rem;padding:.5rem 0">Nenhuma divergência encontrada.</p>'}
    <div style="margin-top:1.5rem;padding-top:.7rem;border-top:1px solid #e5e7eb;font-size:.62rem;color:#aaa;display:flex;justify-content:space-between">
      <span>Sigma Contabilidade · Além da Contabilidade</span>
      <span>Gerado em ${hoje}</span>
    </div>`;

  // Injeta na própria página num overlay de impressão
  let overlay = document.getElementById('print-overlay');
  if(!overlay){
    overlay = document.createElement('div');
    overlay.id = 'print-overlay';
    overlay.style.cssText = 'display:none;position:fixed;inset:0;background:#fff;z-index:9999;overflow:auto;padding:1.5rem 2rem;font-family:Inter,sans-serif';
    document.body.appendChild(overlay);
  }
  overlay.innerHTML = reportHtml;
  overlay.style.display = 'block';
  document.body.style.overflow = 'hidden';

  // Abre o diálogo de impressão
  setTimeout(()=>{
    window.print();
    // Após imprimir, fecha o overlay
    setTimeout(()=>{
      overlay.style.display = 'none';
      document.body.style.overflow = '';
    }, 500);
  }, 400);
}

function tog(el){el.classList.toggle('open')}

function toggleForm(e, btn){
  e.stopPropagation();
  const form = btn.nextElementSibling;
  const isOpen = form.classList.contains('open');
  form.classList.toggle('open', !isOpen);
  btn.textContent = isOpen ? '+ Apontar divergência manual' : '− Cancelar';
}

function calcDiff(inp){
  const form = inp.closest('.div-form');
  const esp = parseFloat(form.querySelector('.f-esp').value) || 0;
  const enc = parseFloat(form.querySelector('.f-enc').value) || 0;
  const diff = form.querySelector('.f-diff');
  if(esp || enc){
    const d = Math.abs(esp - enc);
    diff.textContent = 'R$ ' + d.toLocaleString('pt-BR',{minimumFractionDigits:2,maximumFractionDigits:2});
    diff.style.color = d > 0 ? '#dc2626' : '#059669';
  } else {
    diff.textContent = '—';
    diff.style.color = '';
  }
}

function saveDiv(btn, nomeFunc){
  const form = btn.closest('.div-form');
  const tipo = form.querySelector('.f-tipo').value.trim();
  const desc = form.querySelector('.f-desc').value.trim();
  const esp  = parseFloat(form.querySelector('.f-esp').value) || 0;
  const enc  = parseFloat(form.querySelector('.f-enc').value) || 0;

  if(!tipo && !desc){ alert('Informe ao menos o tipo ou a descrição da divergência.'); return; }

  // Monta texto descritivo
  let descFull = desc || '';
  if(esp || enc){
    const fmtBrl = v => 'R$ ' + v.toLocaleString('pt-BR',{minimumFractionDigits:2,maximumFractionDigits:2});
    const parts = [];
    if(esp) parts.push(`Esperado: ${fmtBrl(esp)}`);
    if(enc) parts.push(`Encontrado: ${fmtBrl(enc)}`);
    if(esp && enc) parts.push(`Diferença: ${fmtBrl(Math.abs(esp-enc))}`);
    descFull += (descFull ? ' | ' : '') + parts.join(' | ');
  }

  // Insere o item de divergência acima do botão
  const divItem = document.createElement('div');
  divItem.className = 'div-item manual';
  divItem.innerHTML = `<div>
    <div class="div-tipo">${tipo || 'Divergência manual'} <span class="manual-tag">manual</span></div>
    ${descFull ? `<div class="div-desc">${descFull}</div>` : ''}
  </div>`;

  const addBtn = form.previousElementSibling;
  addBtn.parentNode.insertBefore(divItem, addBtn);

  // Atualiza badge do card
  const empCard = form.closest('.emp');
  empCard.classList.add('open');
  let badge = empCard.querySelector('.badge');
  const currentDivs = empCard.querySelectorAll('.div-item').length;
  badge.className = 'badge div';
  badge.innerHTML = `⚠ ${currentDivs} divergência${currentDivs>1?'s':''}`;

  // Limpa e fecha form
  form.querySelector('.f-tipo').value = '';
  form.querySelector('.f-desc').value = '';
  form.querySelector('.f-esp').value = '';
  form.querySelector('.f-enc').value = '';
  form.querySelector('.f-diff').textContent = '—';
  form.classList.remove('open');
  addBtn.textContent = '+ Apontar divergência manual';
}
</script>
</body>
</html>
"""

if __name__ == "__main__":
    port  = int(os.environ.get("PORT", 5000))
    debug = os.environ.get("DEBUG", "false").lower() == "true"
    print(f"\n  ✅  Sigma — Conferência de Folha rodando em http://localhost:{port}\n")
    app.run(host="0.0.0.0", port=port, debug=debug)
